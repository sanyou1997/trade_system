"""Excel file reading operations.

Reads inventory, invoice, and daily sales Excel files into plain dicts.
Never modifies the workbooks.
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

import re

from app.excel.config import (
    DATA_END_ROW,
    DATA_START_ROW,
    DAILY_DATA_START_ROW,
    DAILY_HEADER_ROW,
    INV_BRAND_COL,
    INV_COST_COL,
    INV_LISR_COL,
    INV_ORIG_PRICE_COL,
    INV_PATTERN_COL,
    INV_SIZE_COL,
    INV_SUGGESTED_PRICE_COL,
    INV_TYPE_COL,
    INV_PAY_AMOUNT_COL,
    INV_PAY_CUSTOMER_COL,
    INV_PAY_DATE_COL,
    INV_PAY_METHOD_COL,
    INV_SALES_BRAND_COL,
    INV_SALES_CUSTOMER_COL,
    INV_SALES_DATE_COL,
    INV_SALES_DISCOUNT_COL,
    INV_SALES_PAYMENT_COL,
    INV_SALES_PRICE_COL,
    INV_SALES_QTY_COL,
    INV_SALES_SIZE_COL,
    INV_SALES_TOTAL_COL,
    INV_SALES_TYPE_COL,
    INVOICE_LOSS_SHEET,
    INVOICE_OLD_CASH_SHEET,
    INVOICE_OLD_MUKURU_SHEET,
    INVOICE_PAYMENTS_SHEET,
    INVOICE_SALES_SHEET,
    INVOICE_STATS_SHEET,
    OLD_CASH_BRAND_COL,
    OLD_CASH_CUSTOMER_COL,
    OLD_CASH_DATA_START_ROW,
    OLD_CASH_DATE_COL,
    OLD_CASH_NOTE_COL,
    OLD_CASH_PRICE_COL,
    OLD_CASH_QTY_COL,
    OLD_CASH_SIZE_COL,
    OLD_CASH_TOTAL_COL,
    OLD_CASH_TYPE_COL,
    OLD_MUKURU_BRAND_COL,
    OLD_MUKURU_CUSTOMER_COL,
    OLD_MUKURU_DATA_START_ROW,
    OLD_MUKURU_DATE_COL,
    OLD_MUKURU_PRICE_COL,
    OLD_MUKURU_QTY_COL,
    OLD_MUKURU_SIZE_COL,
    OLD_MUKURU_TOTAL_COL,
    OLD_MUKURU_TYPE_COL,
    OLD_PAY_AMOUNT_COL,
    OLD_PAY_CUSTOMER_COL,
    OLD_PAY_DATE_COL,
    STATS_CASH_RATE_ROW,
    STATS_MUKURU_RATE_ROW,
    STATS_RATE_COL,
    get_layout,
    get_sheet_name,
)


def _to_float(value: Any, default: float = 0.0) -> float:
    """Safely convert a cell value to float."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _to_int(value: Any, default: int = 0) -> int:
    """Safely convert a cell value to int."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def _to_str(value: Any) -> str | None:
    """Convert cell value to string or None."""
    if value is None:
        return None
    return str(value).strip() or None


def _excel_date_to_date(value: Any) -> datetime.date | None:
    """Convert Excel serial date, datetime, or string date to Python date.

    Handles:
    - datetime objects
    - date objects
    - Excel serial numbers (int/float)
    - String dates like "2025.9.1" or "2025.9.22" (old format)
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)):
        try:
            return datetime.date.fromordinal(
                datetime.date(1899, 12, 30).toordinal() + int(value)
            )
        except (ValueError, OverflowError):
            return None
    if isinstance(value, str):
        # Try "YYYY.M.D" format (old invoice files)
        m = re.match(r"(\d{4})\.(\d{1,2})\.(\d{1,2})", value.strip())
        if m:
            try:
                return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                return None
    return None


class ExcelReader:
    """Reads data from Excel files without modification."""

    @staticmethod
    def read_inventory(file_path: str | Path, month: int) -> list[dict]:
        """Read all tyre rows from the inventory file for a given month.

        Returns list of dicts with keys: row, size, type, brand, pattern,
        li_sr, tyre_cost, original_price, initial_stock, added_stock,
        daily_sales (dict of day -> qty).
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            sheet_name = get_sheet_name(month)
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. "
                    f"Available: {wb.sheetnames}"
                )
            ws = wb[sheet_name]
            layout = get_layout(month)

            tyres: list[dict] = []
            for row_num in range(DATA_START_ROW, DATA_END_ROW + 1):
                size = _to_str(ws.cell(row_num, INV_SIZE_COL).value)
                if not size:
                    continue
                # Skip summary rows
                if size.lower() in ("total", "total tyre available"):
                    continue

                tyre_type = _to_str(ws.cell(row_num, INV_TYPE_COL).value)
                brand = _to_str(ws.cell(row_num, INV_BRAND_COL).value)
                pattern = _to_str(ws.cell(row_num, INV_PATTERN_COL).value)
                li_sr = _to_str(ws.cell(row_num, INV_LISR_COL).value)
                tyre_cost = _to_float(ws.cell(row_num, INV_COST_COL).value)
                original_price = _to_float(
                    ws.cell(row_num, INV_ORIG_PRICE_COL).value
                )
                suggested_price = _to_float(
                    ws.cell(row_num, INV_SUGGESTED_PRICE_COL).value
                )
                initial_stock = _to_int(
                    ws.cell(row_num, layout.initial_stock_col).value
                )
                added_stock = _to_int(
                    ws.cell(row_num, layout.added_stock_col).value
                )

                # Read daily sales (columns for days 1-31)
                daily_sales: dict[int, int] = {}
                for day in range(1, 32):
                    col = layout.day_to_col(day)
                    qty = _to_int(ws.cell(row_num, col).value)
                    if qty > 0:
                        daily_sales[day] = qty

                tyres.append({
                    "row": row_num,
                    "size": size,
                    "type": tyre_type,
                    "brand": brand,
                    "pattern": pattern,
                    "li_sr": li_sr,
                    "tyre_cost": tyre_cost,
                    "original_price": original_price,
                    "suggested_price": suggested_price,
                    "initial_stock": initial_stock,
                    "added_stock": added_stock,
                    "daily_sales": daily_sales,
                })

            return tyres
        finally:
            wb.close()

    @staticmethod
    def read_exchange_rate(file_path: str | Path, month: int) -> float:
        """Read the exchange rate from the inventory file for a given month."""
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            sheet_name = get_sheet_name(month)
            ws = wb[sheet_name]
            layout = get_layout(month)

            value = ws.cell(
                layout.exchange_rate_row, layout.exchange_rate_col
            ).value
            if value is None:
                raise ValueError(
                    f"Exchange rate not found at "
                    f"{get_column_letter(layout.exchange_rate_col)}"
                    f"{layout.exchange_rate_row}"
                )
            return float(value)
        finally:
            wb.close()

    @staticmethod
    def read_invoice_sales(file_path: str | Path) -> list[dict]:
        """Read sales from the invoice 'Sales Record' sheet.

        Returns list of dicts with keys: date, brand, type, size, qty,
        unit_price, discount, total, payment_method, customer_name.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            ws = wb[INVOICE_SALES_SHEET]
            sales: list[dict] = []

            # Row 1 = headers, row 2+ = data
            for row_num in range(2, ws.max_row + 1):
                date_val = ws.cell(row_num, INV_SALES_DATE_COL).value
                qty = ws.cell(row_num, INV_SALES_QTY_COL).value
                if date_val is None and qty is None:
                    continue  # Skip empty rows

                sale_date = _excel_date_to_date(date_val)
                brand = _to_str(ws.cell(row_num, INV_SALES_BRAND_COL).value)
                sale_type = _to_str(ws.cell(row_num, INV_SALES_TYPE_COL).value)
                size = _to_str(ws.cell(row_num, INV_SALES_SIZE_COL).value)
                quantity = _to_int(qty)
                unit_price = _to_float(
                    ws.cell(row_num, INV_SALES_PRICE_COL).value
                )
                discount = _to_float(
                    ws.cell(row_num, INV_SALES_DISCOUNT_COL).value
                )
                total = _to_float(
                    ws.cell(row_num, INV_SALES_TOTAL_COL).value
                )
                payment_method = _to_str(
                    ws.cell(row_num, INV_SALES_PAYMENT_COL).value
                )
                customer = _to_str(
                    ws.cell(row_num, INV_SALES_CUSTOMER_COL).value
                )

                # Skip rows that look like summaries (e.g., "Total")
                if size and size.lower() == "total":
                    continue

                sales.append({
                    "date": sale_date,
                    "brand": brand,
                    "type": sale_type,
                    "size": size,
                    "qty": quantity,
                    "unit_price": unit_price,
                    "discount": discount,
                    "total": total,
                    "payment_method": payment_method,
                    "customer_name": customer,
                })

            return sales
        finally:
            wb.close()

    @staticmethod
    def read_invoice_payments(file_path: str | Path) -> list[dict]:
        """Read payments from the invoice 'Payment Record' sheet.

        Returns list of dicts with keys: date, customer, payment_method, amount_mwk.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            ws = wb[INVOICE_PAYMENTS_SHEET]
            payments: list[dict] = []

            # Row 1 = headers, row 2+ = data
            for row_num in range(2, ws.max_row + 1):
                amount = ws.cell(row_num, INV_PAY_AMOUNT_COL).value
                if amount is None:
                    continue

                pay_date = _excel_date_to_date(
                    ws.cell(row_num, INV_PAY_DATE_COL).value
                )
                customer = _to_str(
                    ws.cell(row_num, INV_PAY_CUSTOMER_COL).value
                )
                method = _to_str(
                    ws.cell(row_num, INV_PAY_METHOD_COL).value
                )

                payments.append({
                    "date": pay_date,
                    "customer": customer,
                    "payment_method": method,
                    "amount_mwk": _to_float(amount),
                })

            return payments
        finally:
            wb.close()

    @staticmethod
    def read_invoice_losses(file_path: str | Path) -> list[dict]:
        """Read losses from the invoice 'Loss' sheet.

        Returns list of dicts with keys: date, brand, model, config, qty,
        cost, exchanged, refund, total_refund, customer, note.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            ws = wb[INVOICE_LOSS_SHEET]
            losses: list[dict] = []

            # Row 1 = "Invoice" header, Row 2 = column headers, Row 3+ = data
            for row_num in range(3, ws.max_row + 1):
                qty = ws.cell(row_num, 5).value
                if qty is None:
                    continue

                loss_date = _excel_date_to_date(ws.cell(row_num, 1).value)
                brand = _to_str(ws.cell(row_num, 2).value)
                model = _to_str(ws.cell(row_num, 3).value)
                config = _to_str(ws.cell(row_num, 4).value)
                cost = _to_float(ws.cell(row_num, 6).value)
                exchanged = _to_str(ws.cell(row_num, 7).value)
                refund = _to_float(ws.cell(row_num, 8).value)
                total_refund = _to_float(ws.cell(row_num, 9).value)
                customer = _to_str(ws.cell(row_num, 10).value)
                note = _to_str(ws.cell(row_num, 11).value)

                losses.append({
                    "date": loss_date,
                    "brand": brand,
                    "model": model,
                    "config": config,
                    "qty": _to_int(qty),
                    "cost": cost,
                    "exchanged": exchanged,
                    "refund": refund,
                    "total_refund": total_refund,
                    "customer": customer,
                    "note": note,
                })

            return losses
        finally:
            wb.close()

    @staticmethod
    def read_invoice_statistics(file_path: str | Path) -> dict:
        """Read statistics and exchange rates from the invoice 'Statistic' sheet.

        Returns dict with keys: mukuru_rate, cash_rate.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            ws = wb[INVOICE_STATS_SHEET]
            mukuru_rate = _to_float(
                ws.cell(STATS_MUKURU_RATE_ROW, STATS_RATE_COL).value
            )
            cash_rate = _to_float(
                ws.cell(STATS_CASH_RATE_ROW, STATS_RATE_COL).value
            )
            return {
                "mukuru_rate": mukuru_rate,
                "cash_rate": cash_rate,
            }
        finally:
            wb.close()

    @staticmethod
    def read_daily_sales(file_path: str | Path) -> list[dict]:
        """Read sales from a daily sales file ('Tyre Sales DD Mon.xlsx').

        Daily sales files have:
        - Row 1: "Invoice" header
        - Row 2: Column headers (Date, Brand, Type, Size, Qty, ...)
        - Row 3+: Data

        Returns list of dicts with same keys as invoice sales.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            # Daily sales files may have quoted sheet names (e.g. "'Sales Record'")
            sales_sheet = None
            for sn in wb.sheetnames:
                if "sales record" in sn.lower().replace("'", ""):
                    sales_sheet = sn
                    break
            if sales_sheet is None:
                raise ValueError(
                    f"No 'Sales Record' sheet in {file_path}. "
                    f"Available: {wb.sheetnames}"
                )
            ws = wb[sales_sheet]
            sales: list[dict] = []

            for row_num in range(DAILY_DATA_START_ROW, ws.max_row + 1):
                qty = ws.cell(row_num, INV_SALES_QTY_COL).value
                size = _to_str(ws.cell(row_num, INV_SALES_SIZE_COL).value)
                date_val = ws.cell(row_num, INV_SALES_DATE_COL).value

                # Skip empty rows and summary rows
                if qty is None and size is None:
                    continue
                if size and size.lower() == "total":
                    continue
                # Also skip rows where date column says "Total"
                date_str = _to_str(date_val)
                if date_str and date_str.lower() in ("total", "totals"):
                    continue

                sale_date = _excel_date_to_date(date_val)
                brand = _to_str(ws.cell(row_num, INV_SALES_BRAND_COL).value)
                sale_type = _to_str(ws.cell(row_num, INV_SALES_TYPE_COL).value)
                unit_price = _to_float(
                    ws.cell(row_num, INV_SALES_PRICE_COL).value
                )
                discount = _to_float(
                    ws.cell(row_num, INV_SALES_DISCOUNT_COL).value
                )
                total = _to_float(
                    ws.cell(row_num, INV_SALES_TOTAL_COL).value
                )
                payment_method = _to_str(
                    ws.cell(row_num, INV_SALES_PAYMENT_COL).value
                )
                customer = _to_str(
                    ws.cell(row_num, INV_SALES_CUSTOMER_COL).value
                )

                sales.append({
                    "date": sale_date,
                    "brand": brand,
                    "type": sale_type,
                    "size": size,
                    "qty": _to_int(qty),
                    "unit_price": unit_price,
                    "discount": discount,
                    "total": total,
                    "payment_method": payment_method,
                    "customer_name": customer,
                })

            return sales
        finally:
            wb.close()

    @staticmethod
    def detect_invoice_format(file_path: str | Path) -> str:
        """Detect whether an invoice file uses old or new format.

        Returns 'new' if 'Sales Record' sheet exists, 'old' if
        'Cash' or 'Mukuru' sheets exist.
        """
        wb = openpyxl.load_workbook(str(file_path), read_only=True)
        try:
            names = wb.sheetnames
            if INVOICE_SALES_SHEET in names:
                return "new"
            if INVOICE_OLD_CASH_SHEET in names or INVOICE_OLD_MUKURU_SHEET in names:
                return "old"
            raise ValueError(
                f"Unrecognized invoice format. Sheets: {names}. "
                f"Expected '{INVOICE_SALES_SHEET}' (new) or "
                f"'{INVOICE_OLD_CASH_SHEET}'/'{INVOICE_OLD_MUKURU_SHEET}' (old)."
            )
        finally:
            wb.close()

    @staticmethod
    def _read_old_sheet_sales(
        ws: Any,
        payment_method: str,
        *,
        size_col: int,
        type_col: int,
        brand_col: int,
        qty_col: int,
        price_col: int,
        total_col: int,
        customer_col: int,
        date_col: int,
        note_col: int | None,
        data_start_row: int,
    ) -> tuple[list[dict], int]:
        """Read sales rows from an old-format sheet until a gap/summary.

        Returns (sales_list, payment_section_start_row).
        """
        sales: list[dict] = []
        payment_start = ws.max_row + 1  # default: no payment section

        for row_num in range(data_start_row, ws.max_row + 1):
            qty_val = ws.cell(row_num, qty_col).value
            size_val = _to_str(ws.cell(row_num, size_col).value)
            date_val = ws.cell(row_num, date_col).value

            # Detect payment section header: column A contains "Date" text
            date_str = _to_str(date_val)
            if date_str and date_str.lower() == "date":
                payment_start = row_num + 1
                break

            # Skip empty rows
            if qty_val is None and size_val is None and date_val is None:
                continue

            # Skip summary rows
            if size_val and size_val.lower() in ("total", "total quantity"):
                continue
            if date_str and date_str.lower() in (
                "total", "total quantity", "total price", "total price (mkw)",
            ):
                continue

            sale_date = _excel_date_to_date(date_val)
            size = size_val
            brand = _to_str(ws.cell(row_num, brand_col).value)
            sale_type = _to_str(ws.cell(row_num, type_col).value)
            quantity = _to_int(qty_val)
            unit_price = _to_float(ws.cell(row_num, price_col).value)
            total = _to_float(ws.cell(row_num, total_col).value)
            customer = _to_str(ws.cell(row_num, customer_col).value)

            # Old format stores discount in "Note" column as e.g. -0.05
            discount = 0.0
            if note_col is not None:
                note_val = ws.cell(row_num, note_col).value
                if note_val is not None:
                    try:
                        d = float(note_val)
                        # Negative means discount, e.g. -0.05 = 5% off
                        discount = abs(d)
                    except (ValueError, TypeError):
                        pass

            if not size and quantity == 0:
                continue

            sales.append({
                "date": sale_date,
                "brand": brand,
                "type": sale_type,
                "size": size,
                "qty": quantity,
                "unit_price": unit_price,
                "discount": discount,
                "total": total,
                "payment_method": payment_method,
                "customer_name": customer,
            })

        return sales, payment_start

    @staticmethod
    def _read_old_sheet_payments(
        ws: Any,
        payment_method: str,
        start_row: int,
    ) -> list[dict]:
        """Read payments from the embedded payment sub-section of an old sheet."""
        payments: list[dict] = []

        for row_num in range(start_row, ws.max_row + 1):
            amount = ws.cell(row_num, OLD_PAY_AMOUNT_COL).value
            if amount is None:
                continue

            # Skip summary/total rows
            date_str = _to_str(ws.cell(row_num, OLD_PAY_DATE_COL).value)
            if date_str and date_str.lower() in ("total", "totals"):
                continue

            pay_date = _excel_date_to_date(
                ws.cell(row_num, OLD_PAY_DATE_COL).value
            )
            customer = _to_str(
                ws.cell(row_num, OLD_PAY_CUSTOMER_COL).value
            )

            amount_f = _to_float(amount)
            if amount_f <= 0:
                continue

            payments.append({
                "date": pay_date,
                "customer": customer,
                "payment_method": payment_method,
                "amount_mwk": amount_f,
            })

        return payments

    @staticmethod
    def read_invoice_sales_old(file_path: str | Path) -> list[dict]:
        """Read sales from old-format invoice (separate Cash/Mukuru sheets).

        Returns list of dicts with same keys as read_invoice_sales().
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            all_sales: list[dict] = []

            # Read Cash sheet
            if INVOICE_OLD_CASH_SHEET in wb.sheetnames:
                ws = wb[INVOICE_OLD_CASH_SHEET]
                sales, _ = ExcelReader._read_old_sheet_sales(
                    ws, "Cash",
                    size_col=OLD_CASH_SIZE_COL,
                    type_col=OLD_CASH_TYPE_COL,
                    brand_col=OLD_CASH_BRAND_COL,
                    qty_col=OLD_CASH_QTY_COL,
                    price_col=OLD_CASH_PRICE_COL,
                    total_col=OLD_CASH_TOTAL_COL,
                    customer_col=OLD_CASH_CUSTOMER_COL,
                    date_col=OLD_CASH_DATE_COL,
                    note_col=OLD_CASH_NOTE_COL,
                    data_start_row=OLD_CASH_DATA_START_ROW,
                )
                all_sales.extend(sales)

            # Read Mukuru sheet
            if INVOICE_OLD_MUKURU_SHEET in wb.sheetnames:
                ws = wb[INVOICE_OLD_MUKURU_SHEET]
                sales, _ = ExcelReader._read_old_sheet_sales(
                    ws, "Mukuru",
                    size_col=OLD_MUKURU_SIZE_COL,
                    type_col=OLD_MUKURU_TYPE_COL,
                    brand_col=OLD_MUKURU_BRAND_COL,
                    qty_col=OLD_MUKURU_QTY_COL,
                    price_col=OLD_MUKURU_PRICE_COL,
                    total_col=OLD_MUKURU_TOTAL_COL,
                    customer_col=OLD_MUKURU_CUSTOMER_COL,
                    date_col=OLD_MUKURU_DATE_COL,
                    note_col=None,  # Mukuru sheet has no Note column
                    data_start_row=OLD_MUKURU_DATA_START_ROW,
                )
                all_sales.extend(sales)

            return all_sales
        finally:
            wb.close()

    @staticmethod
    def read_invoice_payments_old(file_path: str | Path) -> list[dict]:
        """Read payments from old-format invoice (embedded in Cash/Mukuru sheets).

        Returns list of dicts with same keys as read_invoice_payments().
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            all_payments: list[dict] = []

            # Read Cash sheet payments
            if INVOICE_OLD_CASH_SHEET in wb.sheetnames:
                ws = wb[INVOICE_OLD_CASH_SHEET]
                _, pay_start = ExcelReader._read_old_sheet_sales(
                    ws, "Cash",
                    size_col=OLD_CASH_SIZE_COL,
                    type_col=OLD_CASH_TYPE_COL,
                    brand_col=OLD_CASH_BRAND_COL,
                    qty_col=OLD_CASH_QTY_COL,
                    price_col=OLD_CASH_PRICE_COL,
                    total_col=OLD_CASH_TOTAL_COL,
                    customer_col=OLD_CASH_CUSTOMER_COL,
                    date_col=OLD_CASH_DATE_COL,
                    note_col=OLD_CASH_NOTE_COL,
                    data_start_row=OLD_CASH_DATA_START_ROW,
                )
                payments = ExcelReader._read_old_sheet_payments(
                    ws, "Cash", pay_start,
                )
                all_payments.extend(payments)

            # Read Mukuru sheet payments
            if INVOICE_OLD_MUKURU_SHEET in wb.sheetnames:
                ws = wb[INVOICE_OLD_MUKURU_SHEET]
                _, pay_start = ExcelReader._read_old_sheet_sales(
                    ws, "Mukuru",
                    size_col=OLD_MUKURU_SIZE_COL,
                    type_col=OLD_MUKURU_TYPE_COL,
                    brand_col=OLD_MUKURU_BRAND_COL,
                    qty_col=OLD_MUKURU_QTY_COL,
                    price_col=OLD_MUKURU_PRICE_COL,
                    total_col=OLD_MUKURU_TOTAL_COL,
                    customer_col=OLD_MUKURU_CUSTOMER_COL,
                    date_col=OLD_MUKURU_DATE_COL,
                    note_col=None,
                    data_start_row=OLD_MUKURU_DATA_START_ROW,
                )
                payments = ExcelReader._read_old_sheet_payments(
                    ws, "Mukuru", pay_start,
                )
                all_payments.extend(payments)

            return all_payments
        finally:
            wb.close()

    @staticmethod
    def read_daily_payments(file_path: str | Path) -> list[dict]:
        """Read payments from a daily sales file."""
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            # Daily sales files may have quoted sheet names
            pay_sheet = None
            for sn in wb.sheetnames:
                if "payment record" in sn.lower().replace("'", ""):
                    pay_sheet = sn
                    break
            if pay_sheet is None:
                return []
            ws = wb[pay_sheet]
            payments: list[dict] = []

            for row_num in range(2, ws.max_row + 1):
                amount = ws.cell(row_num, INV_PAY_AMOUNT_COL).value
                if amount is None:
                    continue

                pay_date = _excel_date_to_date(
                    ws.cell(row_num, INV_PAY_DATE_COL).value
                )
                customer = _to_str(
                    ws.cell(row_num, INV_PAY_CUSTOMER_COL).value
                )
                method = _to_str(
                    ws.cell(row_num, INV_PAY_METHOD_COL).value
                )

                payments.append({
                    "date": pay_date,
                    "customer": customer,
                    "payment_method": method,
                    "amount_mwk": _to_float(amount),
                })

            return payments
        finally:
            wb.close()
