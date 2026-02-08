"""Phone Excel file reading operations.

Reads phone inventory, invoice, and daily sales Excel files into plain dicts.
Never modifies the workbooks.
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any

import openpyxl

from app.excel.phone_config import (
    PHONE_CASH_RATE_COL,
    PHONE_CASH_RATE_ROW,
    PHONE_DAILY_DATA_START_ROW,
    PHONE_DATA_MAX_ROW,
    PHONE_DATA_START_ROW,
    PHONE_INV_ADDED_COL,
    PHONE_INV_BRAND_COL,
    PHONE_INV_CASH_PRICE_COL,
    PHONE_INV_CONFIG_COL,
    PHONE_INV_COST_COL,
    PHONE_INV_INITIAL_COL,
    PHONE_INV_MODEL_COL,
    PHONE_INV_MUKURU_PRICE_COL,
    PHONE_INV_NOTE_COL,
    PHONE_INV_ONLINE_PRICE_COL,
    PHONE_INV_PAY_AMOUNT_COL,
    PHONE_INV_PAY_CUSTOMER_COL,
    PHONE_INV_PAY_DATE_COL,
    PHONE_INV_PAY_METHOD_COL,
    PHONE_INV_SALES_BRAND_COL,
    PHONE_INV_SALES_CONFIG_COL,
    PHONE_INV_SALES_CUSTOMER_COL,
    PHONE_INV_SALES_DATE_COL,
    PHONE_INV_SALES_DISCOUNT_COL,
    PHONE_INV_SALES_MODEL_COL,
    PHONE_INV_SALES_PAYMENT_COL,
    PHONE_INV_SALES_PRICE_COL,
    PHONE_INV_SALES_QTY_COL,
    PHONE_INV_SALES_TOTAL_COL,
    PHONE_INV_STATUS_COL,
    PHONE_INVOICE_LOSS_SHEET,
    PHONE_INVOICE_PAYMENTS_SHEET,
    PHONE_INVOICE_SALES_SHEET,
    PHONE_INVOICE_STATS_SHEET,
    PHONE_MUKURU_RATE_COL,
    PHONE_MUKURU_RATE_ROW,
    get_phone_sheet_name,
    phone_day_to_col,
)

# Re-use helper functions from the tyre reader
from app.excel.reader import _excel_date_to_date, _to_float, _to_int, _to_str


class PhoneExcelReader:
    """Reads phone data from Excel files without modification."""

    @staticmethod
    def read_inventory(file_path: str | Path, month: int) -> list[dict]:
        """Read all phone rows from the inventory file for a given month.

        Returns list of dicts with keys: row, brand, model, config, note,
        cost, cash_price, mukuru_price, online_price, status,
        initial_stock, added_stock, daily_sales.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            sheet_name = get_phone_sheet_name(month)
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. "
                    f"Available: {wb.sheetnames}"
                )
            ws = wb[sheet_name]

            phones: list[dict] = []
            for row_num in range(PHONE_DATA_START_ROW, PHONE_DATA_MAX_ROW + 1):
                brand = _to_str(ws.cell(row_num, PHONE_INV_BRAND_COL).value)
                model = _to_str(ws.cell(row_num, PHONE_INV_MODEL_COL).value)
                if not brand and not model:
                    continue  # Skip empty rows
                # Skip summary rows and non-product rows
                if brand and brand.lower() in ("total", "grand total"):
                    continue
                if brand and "return policy" in brand.lower():
                    continue

                config = _to_str(ws.cell(row_num, PHONE_INV_CONFIG_COL).value)
                note = _to_str(ws.cell(row_num, PHONE_INV_NOTE_COL).value)
                cost = _to_float(ws.cell(row_num, PHONE_INV_COST_COL).value)
                cash_price = _to_float(ws.cell(row_num, PHONE_INV_CASH_PRICE_COL).value)
                mukuru_price = _to_float(ws.cell(row_num, PHONE_INV_MUKURU_PRICE_COL).value)
                online_price = _to_float(ws.cell(row_num, PHONE_INV_ONLINE_PRICE_COL).value)
                status = _to_str(ws.cell(row_num, PHONE_INV_STATUS_COL).value)
                initial_stock = _to_int(ws.cell(row_num, PHONE_INV_INITIAL_COL).value)
                added_stock = _to_int(ws.cell(row_num, PHONE_INV_ADDED_COL).value)

                # Read daily sales (columns for days 1-31)
                daily_sales: dict[int, int] = {}
                for day in range(1, 32):
                    col = phone_day_to_col(day)
                    qty = _to_int(ws.cell(row_num, col).value)
                    if qty > 0:
                        daily_sales[day] = qty

                phones.append({
                    "row": row_num,
                    "brand": brand or "",
                    "model": model or "",
                    "config": config or "",
                    "note": note,
                    "cost": cost,
                    "cash_price": cash_price,
                    "mukuru_price": mukuru_price,
                    "online_price": online_price,
                    "status": status,
                    "initial_stock": initial_stock,
                    "added_stock": added_stock,
                    "daily_sales": daily_sales,
                })

            return phones
        finally:
            wb.close()

    @staticmethod
    def read_exchange_rates(
        file_path: str | Path, month: int,
    ) -> dict[str, float]:
        """Read cash and mukuru exchange rates from the phone inventory file.

        Returns dict with keys: cash_rate, mukuru_rate.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            sheet_name = get_phone_sheet_name(month)
            ws = wb[sheet_name]
            cash_rate = _to_float(
                ws.cell(PHONE_CASH_RATE_ROW, PHONE_CASH_RATE_COL).value
            )
            mukuru_rate = _to_float(
                ws.cell(PHONE_MUKURU_RATE_ROW, PHONE_MUKURU_RATE_COL).value
            )
            return {
                "cash_rate": cash_rate,
                "mukuru_rate": mukuru_rate,
            }
        finally:
            wb.close()

    @staticmethod
    def read_invoice_sales(file_path: str | Path) -> list[dict]:
        """Read sales from the phone invoice 'Sales Record' sheet.

        Returns list of dicts with keys: date, brand, model, config, qty,
        unit_price, discount, total, payment_method, customer_name.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            ws = wb[PHONE_INVOICE_SALES_SHEET]
            sales: list[dict] = []

            for row_num in range(2, ws.max_row + 1):
                date_val = ws.cell(row_num, PHONE_INV_SALES_DATE_COL).value
                qty = ws.cell(row_num, PHONE_INV_SALES_QTY_COL).value
                if date_val is None and qty is None:
                    continue

                sale_date = _excel_date_to_date(date_val)
                brand = _to_str(ws.cell(row_num, PHONE_INV_SALES_BRAND_COL).value)
                model = _to_str(ws.cell(row_num, PHONE_INV_SALES_MODEL_COL).value)
                config = _to_str(ws.cell(row_num, PHONE_INV_SALES_CONFIG_COL).value)

                # Skip summary rows
                if brand and brand.lower() == "total":
                    continue

                quantity = _to_int(qty)
                unit_price = _to_float(ws.cell(row_num, PHONE_INV_SALES_PRICE_COL).value)
                discount = _to_float(ws.cell(row_num, PHONE_INV_SALES_DISCOUNT_COL).value)
                total = _to_float(ws.cell(row_num, PHONE_INV_SALES_TOTAL_COL).value)
                payment_method = _to_str(ws.cell(row_num, PHONE_INV_SALES_PAYMENT_COL).value)
                customer = _to_str(ws.cell(row_num, PHONE_INV_SALES_CUSTOMER_COL).value)

                sales.append({
                    "date": sale_date,
                    "brand": brand,
                    "model": model,
                    "config": config,
                    "qty": quantity,
                    "unit_price": unit_price,
                    "discount": discount,
                    "total": total,
                    "payment_method": payment_method,
                    "customer_name": customer,
                })

            return sales
        finally:
            wb.close()

    @staticmethod
    def read_invoice_payments(file_path: str | Path) -> list[dict]:
        """Read payments from the phone invoice 'Payment Record' sheet."""
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            ws = wb[PHONE_INVOICE_PAYMENTS_SHEET]
            payments: list[dict] = []

            for row_num in range(2, ws.max_row + 1):
                amount = ws.cell(row_num, PHONE_INV_PAY_AMOUNT_COL).value
                if amount is None:
                    continue

                pay_date = _excel_date_to_date(
                    ws.cell(row_num, PHONE_INV_PAY_DATE_COL).value
                )
                customer = _to_str(ws.cell(row_num, PHONE_INV_PAY_CUSTOMER_COL).value)
                method = _to_str(ws.cell(row_num, PHONE_INV_PAY_METHOD_COL).value)

                payments.append({
                    "date": pay_date,
                    "customer": customer,
                    "payment_method": method,
                    "amount_mwk": _to_float(amount),
                })

            return payments
        finally:
            wb.close()

    @staticmethod
    def read_invoice_losses(file_path: str | Path) -> list[dict]:
        """Read losses from the phone invoice 'Loss' sheet.

        Returns list of dicts with keys: date, brand, model, config, qty,
        cost, exchanged, refund, total_refund, customer, note.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            ws = wb[PHONE_INVOICE_LOSS_SHEET]
            losses: list[dict] = []

            for row_num in range(3, ws.max_row + 1):
                qty = ws.cell(row_num, 5).value
                if qty is None:
                    continue

                loss_date = _excel_date_to_date(ws.cell(row_num, 1).value)
                brand = _to_str(ws.cell(row_num, 2).value)
                model = _to_str(ws.cell(row_num, 3).value)
                config = _to_str(ws.cell(row_num, 4).value)
                cost = _to_float(ws.cell(row_num, 6).value)
                exchanged = _to_str(ws.cell(row_num, 7).value)
                refund = _to_float(ws.cell(row_num, 8).value)
                total_refund = _to_float(ws.cell(row_num, 9).value)
                customer = _to_str(ws.cell(row_num, 10).value)
                note = _to_str(ws.cell(row_num, 11).value)

                losses.append({
                    "date": loss_date,
                    "brand": brand,
                    "model": model,
                    "config": config,
                    "qty": _to_int(qty),
                    "cost": cost,
                    "exchanged": exchanged,
                    "refund": refund,
                    "total_refund": total_refund,
                    "customer": customer,
                    "note": note,
                })

            return losses
        finally:
            wb.close()

    @staticmethod
    def read_invoice_statistics(file_path: str | Path) -> dict:
        """Read exchange rates from the phone invoice 'Statistic' sheet.

        Uses same layout as tyre invoice: mukuru rate at row 2 col I,
        cash rate at row 3 col I.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            ws = wb[PHONE_INVOICE_STATS_SHEET]
            mukuru_rate = _to_float(ws.cell(2, 9).value)  # I2
            cash_rate = _to_float(ws.cell(3, 9).value)    # I3
            return {
                "mukuru_rate": mukuru_rate,
                "cash_rate": cash_rate,
            }
        finally:
            wb.close()

    @staticmethod
    def read_daily_sales(file_path: str | Path) -> list[dict]:
        """Read sales from a daily phone sales file.

        Daily sales files have:
        - Row 1: "Invoice" header
        - Row 2: Column headers
        - Row 3+: Data

        Returns list of dicts with same keys as invoice sales.
        """
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            sales_sheet = None
            for sn in wb.sheetnames:
                if "sales record" in sn.lower().replace("'", ""):
                    sales_sheet = sn
                    break
            if sales_sheet is None:
                raise ValueError(
                    f"No 'Sales Record' sheet in {file_path}. "
                    f"Available: {wb.sheetnames}"
                )
            ws = wb[sales_sheet]
            sales: list[dict] = []

            for row_num in range(PHONE_DAILY_DATA_START_ROW, ws.max_row + 1):
                qty = ws.cell(row_num, PHONE_INV_SALES_QTY_COL).value
                brand = _to_str(ws.cell(row_num, PHONE_INV_SALES_BRAND_COL).value)
                date_val = ws.cell(row_num, PHONE_INV_SALES_DATE_COL).value

                if qty is None and brand is None:
                    continue
                date_str = _to_str(date_val)
                if date_str and date_str.lower() in ("total", "totals"):
                    continue
                if brand and brand.lower() == "total":
                    continue

                sale_date = _excel_date_to_date(date_val)
                model = _to_str(ws.cell(row_num, PHONE_INV_SALES_MODEL_COL).value)
                config = _to_str(ws.cell(row_num, PHONE_INV_SALES_CONFIG_COL).value)
                unit_price = _to_float(ws.cell(row_num, PHONE_INV_SALES_PRICE_COL).value)
                discount = _to_float(ws.cell(row_num, PHONE_INV_SALES_DISCOUNT_COL).value)
                total = _to_float(ws.cell(row_num, PHONE_INV_SALES_TOTAL_COL).value)
                payment_method = _to_str(ws.cell(row_num, PHONE_INV_SALES_PAYMENT_COL).value)
                customer = _to_str(ws.cell(row_num, PHONE_INV_SALES_CUSTOMER_COL).value)

                sales.append({
                    "date": sale_date,
                    "brand": brand,
                    "model": model,
                    "config": config,
                    "qty": _to_int(qty),
                    "unit_price": unit_price,
                    "discount": discount,
                    "total": total,
                    "payment_method": payment_method,
                    "customer_name": customer,
                })

            return sales
        finally:
            wb.close()

    @staticmethod
    def read_daily_payments(file_path: str | Path) -> list[dict]:
        """Read payments from a daily phone sales file."""
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        try:
            pay_sheet = None
            for sn in wb.sheetnames:
                if "payment record" in sn.lower().replace("'", ""):
                    pay_sheet = sn
                    break
            if pay_sheet is None:
                return []
            ws = wb[pay_sheet]
            payments: list[dict] = []

            for row_num in range(2, ws.max_row + 1):
                amount = ws.cell(row_num, PHONE_INV_PAY_AMOUNT_COL).value
                if amount is None:
                    continue

                pay_date = _excel_date_to_date(
                    ws.cell(row_num, PHONE_INV_PAY_DATE_COL).value
                )
                customer = _to_str(ws.cell(row_num, PHONE_INV_PAY_CUSTOMER_COL).value)
                method = _to_str(ws.cell(row_num, PHONE_INV_PAY_METHOD_COL).value)

                payments.append({
                    "date": pay_date,
                    "customer": customer,
                    "payment_method": method,
                    "amount_mwk": _to_float(amount),
                })

            return payments
        finally:
            wb.close()
