"""Phone Excel file writing operations with formula protection.

CRITICAL: Never overwrites phone formula columns (E, G, H, I, J).
Always creates a backup before any write operation.
"""

from __future__ import annotations

import datetime
import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from app.excel.phone_config import (
    PHONE_DATA_MAX_ROW,
    PHONE_DATA_START_ROW,
    PHONE_FORMULA_COLUMNS,
    PHONE_INV_ADDED_COL,
    PHONE_INV_DAILY_END_COL,
    PHONE_INV_DAILY_START_COL,
    PHONE_INV_INITIAL_COL,
    PHONE_INV_PAY_AMOUNT_COL,
    PHONE_INV_PAY_CUSTOMER_COL,
    PHONE_INV_PAY_DATE_COL,
    PHONE_INV_PAY_METHOD_COL,
    PHONE_INV_SALES_BRAND_COL,
    PHONE_INV_SALES_CONFIG_COL,
    PHONE_INV_SALES_CUSTOMER_COL,
    PHONE_INV_SALES_DATE_COL,
    PHONE_INV_SALES_DISCOUNT_COL,
    PHONE_INV_SALES_MODEL_COL,
    PHONE_INV_SALES_PAYMENT_COL,
    PHONE_INV_SALES_PRICE_COL,
    PHONE_INV_SALES_QTY_COL,
    PHONE_INV_SALES_TOTAL_COL,
    PHONE_INVOICE_PAYMENTS_SHEET,
    PHONE_INVOICE_SALES_SHEET,
    get_phone_sheet_name,
    phone_day_to_col,
)


def _phone_safe_write(ws: object, row: int, col: int, value: object) -> None:
    """Write a value to a cell, RAISING if the column contains formulas.

    Raises:
        ValueError: If col is in PHONE_FORMULA_COLUMNS (E, G, H, I, J).
    """
    if col in PHONE_FORMULA_COLUMNS:
        col_letter = get_column_letter(col)
        raise ValueError(
            f"REFUSED: Column {col_letter} (index {col}) is a phone formula column. "
            f"Writing to formula columns {sorted(PHONE_FORMULA_COLUMNS)} is forbidden."
        )
    ws.cell(row=row, column=col, value=value)


def _create_backup(file_path: Path) -> Path:
    """Create a timestamped backup of the file before modification."""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = file_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{file_path.stem}_{timestamp}{file_path.suffix}"
    shutil.copy2(str(file_path), str(backup_path))
    return backup_path


class PhoneExcelWriter:
    """Writes phone data to Excel files with formula protection."""

    @staticmethod
    def ensure_month_sheet(file_path: str | Path, month: int) -> bool:
        """Ensure the month sheet exists in the phone inventory file.

        Returns True if a new sheet was created, False if it already existed.
        """
        path = Path(file_path)
        target_name = get_phone_sheet_name(month)
        source_name = get_phone_sheet_name(1)

        wb = openpyxl.load_workbook(str(path))
        try:
            if target_name in wb.sheetnames:
                return False

            if source_name not in wb.sheetnames:
                source_name = wb.sheetnames[0]

            source_ws = wb[source_name]
            target_ws = wb.copy_worksheet(source_ws)
            target_ws.title = target_name

            # Clear daily sales and stock columns
            for row in range(PHONE_DATA_START_ROW, PHONE_DATA_MAX_ROW + 1):
                for col in range(PHONE_INV_DAILY_START_COL, PHONE_INV_DAILY_END_COL + 1):
                    target_ws.cell(row=row, column=col, value=None)
                target_ws.cell(row=row, column=PHONE_INV_INITIAL_COL, value=None)
                target_ws.cell(row=row, column=PHONE_INV_ADDED_COL, value=None)

            wb.save(str(path))
            return True
        finally:
            wb.close()

    @staticmethod
    def create_invoice_file(file_path: str | Path) -> None:
        """Create a new phone invoice Excel file with standard sheet structure."""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # Sales Record sheet
        ws_sales = wb.active
        ws_sales.title = PHONE_INVOICE_SALES_SHEET
        sales_headers = [
            "Date", "Brand", "Model", "Config", "Qty",
            "Unit Price", "Discount", "Total", "Payment Method",
            "Customer Name",
        ]
        for col_idx, header in enumerate(sales_headers, 1):
            ws_sales.cell(row=1, column=col_idx, value=header)

        # Payment Record sheet
        ws_pay = wb.create_sheet(PHONE_INVOICE_PAYMENTS_SHEET)
        pay_headers = ["Date", "Customer", "Payment Method", "MWK"]
        for col_idx, header in enumerate(pay_headers, 1):
            ws_pay.cell(row=1, column=col_idx, value=header)

        # Loss sheet
        ws_loss = wb.create_sheet("Loss")
        loss_headers = [
            "Date", "Brand", "Model", "Config", "Qty",
            "Cost", "Exchanged", "Refund per pc", "Total Refund",
            "Customer", "Note",
        ]
        for col_idx, header in enumerate(loss_headers, 1):
            ws_loss.cell(row=1, column=col_idx, value=header)

        # Statistic sheet
        ws_stats = wb.create_sheet("Statistic")
        ws_stats.cell(row=1, column=1, value="Statistic")

        # Broken Stock sheet
        wb.create_sheet("Broken Stock")

        wb.save(str(path))
        wb.close()

    @staticmethod
    def export_inventory_batch(
        file_path: str | Path,
        month: int,
        stock_data: list[dict],
        sales_by_day: dict[int, list[dict]],
    ) -> int:
        """Write all phone inventory data in a single file open/save cycle.

        Clears daily sales and stock columns first, then writes all data.

        Args:
            file_path: Path to the phone inventory Excel file.
            month: Month number (1-12).
            stock_data: List of dicts with 'row', 'initial_stock', 'added_stock'.
            sales_by_day: Dict mapping day number to list of
                dicts with 'row' and 'qty' keys.

        Returns:
            Number of records written.
        """
        path = Path(file_path)
        _create_backup(path)

        sheet_name = get_phone_sheet_name(month)
        records = 0

        wb = openpyxl.load_workbook(str(path))
        try:
            ws = wb[sheet_name]

            # Clear daily sales and stock columns
            for row in range(PHONE_DATA_START_ROW, PHONE_DATA_MAX_ROW + 1):
                ws.cell(row=row, column=PHONE_INV_INITIAL_COL, value=None)
                ws.cell(row=row, column=PHONE_INV_ADDED_COL, value=None)
                for col in range(PHONE_INV_DAILY_START_COL, PHONE_INV_DAILY_END_COL + 1):
                    ws.cell(row=row, column=col, value=None)

            # Write stock levels
            for entry in stock_data:
                row = entry["row"]
                init = entry.get("initial_stock", 0)
                added = entry.get("added_stock", 0)
                if init:
                    _phone_safe_write(ws, row, PHONE_INV_INITIAL_COL, init)
                if added:
                    _phone_safe_write(ws, row, PHONE_INV_ADDED_COL, added)

            # Write daily sales
            for day, sales in sorted(sales_by_day.items()):
                col = phone_day_to_col(day)
                for sale in sales:
                    _phone_safe_write(ws, sale["row"], col, sale["qty"] or None)
                    records += 1

            wb.save(str(path))
            return records
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()

    @staticmethod
    def export_invoice_batch(
        file_path: str | Path,
        sales: list[dict],
        payments: list[dict],
    ) -> tuple[int, int]:
        """Write all phone sales and payments to invoice file in one open/save.

        Clears existing data (keeps headers), then writes all records.

        Returns:
            Tuple of (sales_written, payments_written).
        """
        path = Path(file_path)
        _create_backup(path)

        wb = openpyxl.load_workbook(str(path))
        try:
            # --- Sales Record ---
            ws_sales = wb[PHONE_INVOICE_SALES_SHEET]
            if ws_sales.max_row > 1:
                ws_sales.delete_rows(2, ws_sales.max_row - 1)

            for i, sale in enumerate(sales):
                row = i + 2
                sale_date = sale.get("date")
                if isinstance(sale_date, datetime.date):
                    sale_date = datetime.datetime(
                        sale_date.year, sale_date.month, sale_date.day
                    )

                ws_sales.cell(row, PHONE_INV_SALES_DATE_COL, sale_date)
                ws_sales.cell(row, PHONE_INV_SALES_BRAND_COL, sale.get("brand"))
                ws_sales.cell(row, PHONE_INV_SALES_MODEL_COL, sale.get("model"))
                ws_sales.cell(row, PHONE_INV_SALES_CONFIG_COL, sale.get("config"))
                ws_sales.cell(row, PHONE_INV_SALES_QTY_COL, sale.get("qty"))
                ws_sales.cell(row, PHONE_INV_SALES_PRICE_COL, sale.get("unit_price"))
                ws_sales.cell(row, PHONE_INV_SALES_DISCOUNT_COL, sale.get("discount"))
                # Total formula: =E*F*(1-G)
                e = get_column_letter(PHONE_INV_SALES_QTY_COL)
                f = get_column_letter(PHONE_INV_SALES_PRICE_COL)
                g = get_column_letter(PHONE_INV_SALES_DISCOUNT_COL)
                ws_sales.cell(
                    row, PHONE_INV_SALES_TOTAL_COL,
                    f"={e}{row}*{f}{row}*(1-{g}{row})",
                )
                ws_sales.cell(row, PHONE_INV_SALES_PAYMENT_COL, sale.get("payment_method"))
                ws_sales.cell(row, PHONE_INV_SALES_CUSTOMER_COL, sale.get("customer_name"))

            # --- Payment Record ---
            ws_pay = wb[PHONE_INVOICE_PAYMENTS_SHEET]
            if ws_pay.max_row > 1:
                ws_pay.delete_rows(2, ws_pay.max_row - 1)

            for i, payment in enumerate(payments):
                row = i + 2
                pay_date = payment.get("date")
                if isinstance(pay_date, datetime.date):
                    pay_date = datetime.datetime(
                        pay_date.year, pay_date.month, pay_date.day
                    )

                ws_pay.cell(row, PHONE_INV_PAY_DATE_COL, pay_date)
                ws_pay.cell(row, PHONE_INV_PAY_CUSTOMER_COL, payment.get("customer"))
                ws_pay.cell(row, PHONE_INV_PAY_METHOD_COL, payment.get("payment_method"))
                ws_pay.cell(row, PHONE_INV_PAY_AMOUNT_COL, payment.get("amount_mwk"))

            wb.save(str(path))
            return len(sales), len(payments)
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()
