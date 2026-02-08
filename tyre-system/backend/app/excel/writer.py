"""Excel file writing operations with formula protection.

CRITICAL: Never overwrites formula columns (G, H, I, J, K).
Always creates a backup before any write operation.
"""

from __future__ import annotations

import datetime
import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from app.excel.config import (
    DATA_END_ROW,
    DATA_START_ROW,
    FORMULA_COLUMNS,
    INV_PAY_AMOUNT_COL,
    INV_PAY_CUSTOMER_COL,
    INV_PAY_DATE_COL,
    INV_PAY_METHOD_COL,
    INV_SALES_BRAND_COL,
    INV_SALES_CUSTOMER_COL,
    INV_SALES_DATE_COL,
    INV_SALES_DISCOUNT_COL,
    INV_SALES_PAYMENT_COL,
    INV_SALES_PRICE_COL,
    INV_SALES_QTY_COL,
    INV_SALES_SIZE_COL,
    INV_SALES_TOTAL_COL,
    INV_SALES_TYPE_COL,
    INVOICE_PAYMENTS_SHEET,
    INVOICE_SALES_SHEET,
    get_layout,
    get_sheet_name,
)


def _safe_write(ws: object, row: int, col: int, value: object) -> None:
    """Write a value to a cell, RAISING if the column contains formulas.

    Args:
        ws: The openpyxl worksheet.
        row: Row number (1-indexed).
        col: Column number (1-indexed).
        value: Value to write.

    Raises:
        ValueError: If col is in FORMULA_COLUMNS (G, H, I, J, K).
    """
    if col in FORMULA_COLUMNS:
        col_letter = get_column_letter(col)
        raise ValueError(
            f"REFUSED: Column {col_letter} (index {col}) is a formula column. "
            f"Writing to formula columns {sorted(FORMULA_COLUMNS)} is forbidden."
        )
    ws.cell(row=row, column=col, value=value)


def _create_backup(file_path: Path) -> Path:
    """Create a timestamped backup of the file before modification.

    Returns the backup file path.
    """
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = file_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{file_path.stem}_{timestamp}{file_path.suffix}"
    shutil.copy2(str(file_path), str(backup_path))
    return backup_path


class ExcelWriter:
    """Writes data to Excel files with formula protection."""

    @staticmethod
    def ensure_month_sheet(file_path: str | Path, month: int) -> bool:
        """Ensure the month sheet exists in the inventory file.

        If the sheet doesn't exist, copies the structure from month 1's
        sheet (headers, formulas, formatting) and clears daily sales data.

        Returns True if a new sheet was created, False if it already existed.
        """
        path = Path(file_path)
        target_name = get_sheet_name(month)
        source_name = get_sheet_name(1)  # Copy from month 1

        wb = openpyxl.load_workbook(str(path))
        try:
            if target_name in wb.sheetnames:
                return False

            if source_name not in wb.sheetnames:
                # Fall back to the first available sheet
                source_name = wb.sheetnames[0]

            source_ws = wb[source_name]
            target_ws = wb.copy_worksheet(source_ws)
            target_ws.title = target_name

            # Clear daily sales columns (O-AS or P-AT) but keep
            # headers, static data, and formulas in formula columns
            layout = get_layout(month)
            for row in range(DATA_START_ROW, DATA_END_ROW + 1):
                for col in range(layout.daily_start_col, layout.daily_end_col + 1):
                    target_ws.cell(row=row, column=col, value=None)
                # Clear initial stock and added stock too (new month = fresh)
                target_ws.cell(
                    row=row, column=layout.initial_stock_col, value=None
                )
                target_ws.cell(
                    row=row, column=layout.added_stock_col, value=None
                )

            wb.save(str(path))
            return True
        finally:
            wb.close()

    @staticmethod
    def create_invoice_file(file_path: str | Path) -> None:
        """Create a new invoice Excel file with standard sheet structure.

        Creates sheets: Sales Record, Payment Record, Loss, Statistic,
        Broken Stock — with column headers.
        """
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # Sales Record sheet
        ws_sales = wb.active
        ws_sales.title = INVOICE_SALES_SHEET
        sales_headers = [
            "Date", "Brand", "Type", "Size", "Qty",
            "Unit Price", "Discount", "Total", "Payment Method",
            "Customer Name",
        ]
        for col_idx, header in enumerate(sales_headers, 1):
            ws_sales.cell(row=1, column=col_idx, value=header)

        # Payment Record sheet
        ws_pay = wb.create_sheet(INVOICE_PAYMENTS_SHEET)
        pay_headers = ["Date", "Customer", "Payment Method", "MWK"]
        for col_idx, header in enumerate(pay_headers, 1):
            ws_pay.cell(row=1, column=col_idx, value=header)

        # Loss sheet
        ws_loss = wb.create_sheet("Loss")
        loss_headers = [
            "Date", "Brand", "Model", "Config", "Qty",
            "Exchanged", "Refund per pc", "Total Refund", "Note",
        ]
        for col_idx, header in enumerate(loss_headers, 1):
            ws_loss.cell(row=1, column=col_idx, value=header)

        # Statistic sheet
        ws_stats = wb.create_sheet("Statistic")
        ws_stats.cell(row=1, column=1, value="Statistic")

        # Broken Stock sheet
        wb.create_sheet("Broken Stock")

        wb.save(str(path))
        wb.close()

    @staticmethod
    def write_daily_sales(
        file_path: str | Path,
        month: int,
        day: int,
        sales: list[dict],
    ) -> None:
        """Write daily sales quantities to the inventory file.

        Each dict in sales should have: row (int), qty (int).
        Writes to the column for the given day.

        Args:
            file_path: Path to the inventory Excel file.
            month: Month number (1-12).
            day: Day of month (1-31).
            sales: List of dicts with 'row' and 'qty' keys.
        """
        path = Path(file_path)
        _create_backup(path)

        layout = get_layout(month)
        col = layout.day_to_col(day)
        sheet_name = get_sheet_name(month)

        wb = openpyxl.load_workbook(str(path))
        try:
            ws = wb[sheet_name]
            for sale in sales:
                row = sale["row"]
                qty = sale["qty"]
                _safe_write(ws, row, col, qty if qty else None)
            wb.save(str(path))
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()

    @staticmethod
    def write_initial_stock(
        file_path: str | Path,
        month: int,
        tyre_row: int,
        value: int,
    ) -> None:
        """Write initial stock value for a tyre in the inventory file.

        Args:
            file_path: Path to the inventory Excel file.
            month: Month number (1-12).
            tyre_row: Row number of the tyre (1-indexed).
            value: Initial stock value.
        """
        path = Path(file_path)
        _create_backup(path)

        layout = get_layout(month)
        sheet_name = get_sheet_name(month)

        wb = openpyxl.load_workbook(str(path))
        try:
            ws = wb[sheet_name]
            _safe_write(ws, tyre_row, layout.initial_stock_col, value)
            wb.save(str(path))
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()

    @staticmethod
    def write_added_stock(
        file_path: str | Path,
        month: int,
        tyre_row: int,
        value: int,
    ) -> None:
        """Write added stock value for a tyre in the inventory file.

        Args:
            file_path: Path to the inventory Excel file.
            month: Month number (1-12).
            tyre_row: Row number of the tyre (1-indexed).
            value: Added stock value.
        """
        path = Path(file_path)
        _create_backup(path)

        layout = get_layout(month)
        sheet_name = get_sheet_name(month)

        wb = openpyxl.load_workbook(str(path))
        try:
            ws = wb[sheet_name]
            _safe_write(ws, tyre_row, layout.added_stock_col, value)
            wb.save(str(path))
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()

    @staticmethod
    def export_inventory_batch(
        file_path: str | Path,
        month: int,
        stock_data: list[dict],
        sales_by_day: dict[int, list[dict]],
    ) -> int:
        """Write all inventory data in a single file open/save cycle.

        Clears daily sales and stock columns first, then writes all data.
        This avoids the N-times-open problem and ensures stale data is removed.

        Args:
            file_path: Path to the inventory Excel file.
            month: Month number (1-12).
            stock_data: List of dicts with 'row', 'initial_stock', 'added_stock'.
            sales_by_day: Dict mapping day number to list of
                dicts with 'row' and 'qty' keys.

        Returns:
            Number of records written.
        """
        path = Path(file_path)
        _create_backup(path)

        layout = get_layout(month)
        sheet_name = get_sheet_name(month)
        records = 0

        wb = openpyxl.load_workbook(str(path))
        try:
            ws = wb[sheet_name]

            # Clear daily sales columns and stock columns first
            for row in range(DATA_START_ROW, DATA_END_ROW + 1):
                ws.cell(row=row, column=layout.initial_stock_col, value=None)
                ws.cell(row=row, column=layout.added_stock_col, value=None)
                for col in range(layout.daily_start_col, layout.daily_end_col + 1):
                    ws.cell(row=row, column=col, value=None)

            # Write stock levels
            for entry in stock_data:
                row = entry["row"]
                init = entry.get("initial_stock", 0)
                added = entry.get("added_stock", 0)
                if init:
                    _safe_write(ws, row, layout.initial_stock_col, init)
                if added:
                    _safe_write(ws, row, layout.added_stock_col, added)

            # Write daily sales
            for day, sales in sorted(sales_by_day.items()):
                col = layout.day_to_col(day)
                for sale in sales:
                    _safe_write(ws, sale["row"], col, sale["qty"] or None)
                    records += 1

            wb.save(str(path))
            return records
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()

    @staticmethod
    def export_invoice_batch(
        file_path: str | Path,
        sales: list[dict],
        payments: list[dict],
    ) -> tuple[int, int]:
        """Write all sales and payments to invoice file in one open/save.

        Clears existing Sales Record and Payment Record data (keeps headers),
        then writes all records. This makes each export produce a complete file.

        Args:
            file_path: Path to invoice Excel file.
            sales: List of sale dicts.
            payments: List of payment dicts.

        Returns:
            Tuple of (sales_written, payments_written).
        """
        path = Path(file_path)
        _create_backup(path)

        wb = openpyxl.load_workbook(str(path))
        try:
            # --- Sales Record ---
            ws_sales = wb[INVOICE_SALES_SHEET]

            # Delete all data rows (keep header row 1)
            if ws_sales.max_row > 1:
                ws_sales.delete_rows(2, ws_sales.max_row - 1)

            # Write all sales
            for i, sale in enumerate(sales):
                row = i + 2  # start at row 2

                sale_date = sale.get("date")
                if isinstance(sale_date, datetime.date):
                    sale_date = datetime.datetime(
                        sale_date.year, sale_date.month, sale_date.day
                    )

                ws_sales.cell(row, INV_SALES_DATE_COL, sale_date)
                ws_sales.cell(row, INV_SALES_BRAND_COL, sale.get("brand"))
                ws_sales.cell(row, INV_SALES_TYPE_COL, sale.get("type"))
                ws_sales.cell(row, INV_SALES_SIZE_COL, sale.get("size"))
                ws_sales.cell(row, INV_SALES_QTY_COL, sale.get("qty"))
                ws_sales.cell(row, INV_SALES_PRICE_COL, sale.get("unit_price"))
                ws_sales.cell(row, INV_SALES_DISCOUNT_COL, sale.get("discount"))
                # Total formula: =E*F*(1-G)
                e = get_column_letter(INV_SALES_QTY_COL)
                f = get_column_letter(INV_SALES_PRICE_COL)
                g = get_column_letter(INV_SALES_DISCOUNT_COL)
                ws_sales.cell(
                    row, INV_SALES_TOTAL_COL,
                    f"={e}{row}*{f}{row}*(1-{g}{row})",
                )
                ws_sales.cell(row, INV_SALES_PAYMENT_COL, sale.get("payment_method"))
                ws_sales.cell(row, INV_SALES_CUSTOMER_COL, sale.get("customer_name"))

            # --- Payment Record ---
            ws_pay = wb[INVOICE_PAYMENTS_SHEET]

            # Delete all data rows (keep header row 1)
            if ws_pay.max_row > 1:
                ws_pay.delete_rows(2, ws_pay.max_row - 1)

            # Write all payments
            for i, payment in enumerate(payments):
                row = i + 2

                pay_date = payment.get("date")
                if isinstance(pay_date, datetime.date):
                    pay_date = datetime.datetime(
                        pay_date.year, pay_date.month, pay_date.day
                    )

                ws_pay.cell(row, INV_PAY_DATE_COL, pay_date)
                ws_pay.cell(row, INV_PAY_CUSTOMER_COL, payment.get("customer"))
                ws_pay.cell(row, INV_PAY_METHOD_COL, payment.get("payment_method"))
                ws_pay.cell(row, INV_PAY_AMOUNT_COL, payment.get("amount_mwk"))

            wb.save(str(path))
            return len(sales), len(payments)
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()

    @staticmethod
    def append_invoice_sale(
        file_path: str | Path,
        sale: dict,
    ) -> None:
        """Append a sale record to the invoice 'Sales Record' sheet.

        Dict should have: date, brand, type, size, qty, unit_price,
        discount, payment_method, customer_name.
        The total column (H) uses a formula, so we don't write it.
        """
        path = Path(file_path)
        _create_backup(path)

        wb = openpyxl.load_workbook(str(path))
        try:
            ws = wb[INVOICE_SALES_SHEET]
            # Find next empty row
            next_row = ws.max_row + 1

            # Convert date to Excel serial number if it's a date object
            sale_date = sale.get("date")
            if isinstance(sale_date, datetime.date):
                sale_date = (
                    sale_date - datetime.date(1899, 12, 30)
                ).days

            ws.cell(next_row, INV_SALES_DATE_COL, sale_date)
            ws.cell(next_row, INV_SALES_BRAND_COL, sale.get("brand"))
            ws.cell(next_row, INV_SALES_TYPE_COL, sale.get("type"))
            ws.cell(next_row, INV_SALES_SIZE_COL, sale.get("size"))
            ws.cell(next_row, INV_SALES_QTY_COL, sale.get("qty"))
            ws.cell(next_row, INV_SALES_PRICE_COL, sale.get("unit_price"))
            ws.cell(next_row, INV_SALES_DISCOUNT_COL, sale.get("discount"))
            # Column H (total) = formula: =E*F*(1-G)
            e_col = get_column_letter(INV_SALES_QTY_COL)
            f_col = get_column_letter(INV_SALES_PRICE_COL)
            g_col = get_column_letter(INV_SALES_DISCOUNT_COL)
            ws.cell(
                next_row,
                INV_SALES_TOTAL_COL,
                f"={e_col}{next_row}*{f_col}{next_row}*(1-{g_col}{next_row})",
            )
            ws.cell(
                next_row, INV_SALES_PAYMENT_COL, sale.get("payment_method")
            )
            ws.cell(
                next_row, INV_SALES_CUSTOMER_COL, sale.get("customer_name")
            )

            wb.save(str(path))
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()

    @staticmethod
    def append_invoice_payment(
        file_path: str | Path,
        payment: dict,
    ) -> None:
        """Append a payment record to the invoice 'Payment Record' sheet.

        Dict should have: date, customer, payment_method, amount_mwk.
        """
        path = Path(file_path)
        _create_backup(path)

        wb = openpyxl.load_workbook(str(path))
        try:
            ws = wb[INVOICE_PAYMENTS_SHEET]
            next_row = ws.max_row + 1

            pay_date = payment.get("date")
            if isinstance(pay_date, datetime.date):
                pay_date = datetime.datetime(
                    pay_date.year, pay_date.month, pay_date.day
                )

            ws.cell(next_row, INV_PAY_DATE_COL, pay_date)
            ws.cell(next_row, INV_PAY_CUSTOMER_COL, payment.get("customer"))
            ws.cell(
                next_row, INV_PAY_METHOD_COL, payment.get("payment_method")
            )
            ws.cell(next_row, INV_PAY_AMOUNT_COL, payment.get("amount_mwk"))

            wb.save(str(path))
        except Exception:
            wb.close()
            raise
        finally:
            wb.close()
