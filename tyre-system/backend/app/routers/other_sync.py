"""Other products sync router - import from simple Excel files."""

from __future__ import annotations

import datetime
import hashlib
from pathlib import Path

import shutil
import tempfile

from fastapi import APIRouter, Depends, File, Query, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.other_product import OtherProduct
from app.models.other_inventory import OtherInventoryPeriod
from app.models.sync_log import SyncDirection, SyncLog, SyncStatus
from app.schemas.common import ApiResponse

router = APIRouter(prefix="/other-sync", tags=["other-sync"])


def _save_upload(upload: UploadFile) -> Path:
    """Save an uploaded file to a temp file and return the path."""
    suffix = Path(upload.filename or "upload.xlsx").suffix or ".xlsx"
    fd, tmp_path = tempfile.mkstemp(suffix=suffix)
    try:
        with open(fd, "wb") as f:
            shutil.copyfileobj(upload.file, f)
    except Exception:
        Path(tmp_path).unlink(missing_ok=True)
        raise
    return Path(tmp_path)


def _compute_file_hash(file_path: str) -> str:
    """Compute MD5 hash of a file."""
    h = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _normalize(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().lower()


async def _log_sync(
    db: AsyncSession,
    file_path: str,
    direction: SyncDirection,
    status: SyncStatus,
    records: int = 0,
    error: str | None = None,
    file_hash: str | None = None,
) -> SyncLog:
    log = SyncLog(
        file_path=file_path,
        direction=direction,
        status=status,
        records_processed=records,
        error_message=error,
        file_hash=file_hash,
    )
    db.add(log)
    await db.commit()
    return log


@router.post("/import/inventory")
async def import_other_inventory(
    file: UploadFile = File(..., description="Other products Excel file (.xlsx)"),
    month: int = Query(default=None, ge=1, le=12),
    year: int = Query(default=None),
    db: AsyncSession = Depends(get_db),
) -> ApiResponse[dict]:
    """Import other products + stock from an uploaded Excel file.

    Expected format: 3 columns (A=Name, B=Unit Price MWK, C=QTY).
    Creates product records if they don't exist (matched by normalized name).
    Updates inventory periods for the given month.
    """
    import openpyxl

    now = datetime.date.today()
    year = year or now.year
    month = month or now.month

    file_path = _save_upload(file)
    original_name = file.filename or "other_inventory.xlsx"

    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        ws = wb.active

        imported = 0
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=1).value
            price = ws.cell(row=row_idx, column=2).value
            qty = ws.cell(row=row_idx, column=3).value

            if not name:
                continue

            name = str(name).strip()
            price = float(price) if price else 0.0
            qty = int(float(qty)) if qty else 0

            # Match by normalized name
            result = await db.execute(
                select(OtherProduct).where(
                    OtherProduct.name == name,
                )
            )
            product = result.scalar_one_or_none()

            # Fallback: case-insensitive match
            if product is None:
                all_result = await db.execute(select(OtherProduct))
                for p in all_result.scalars().all():
                    if _normalize(p.name) == _normalize(name):
                        product = p
                        break

            if product is None:
                product = OtherProduct(
                    name=name,
                    suggested_price=price,
                    cost=0.0,
                    excel_row=row_idx,
                )
                db.add(product)
                await db.flush()
            else:
                product.suggested_price = price
                product.excel_row = row_idx

            # Upsert inventory period
            inv_result = await db.execute(
                select(OtherInventoryPeriod).where(
                    OtherInventoryPeriod.other_product_id == product.id,
                    OtherInventoryPeriod.year == year,
                    OtherInventoryPeriod.month == month,
                )
            )
            inv = inv_result.scalar_one_or_none()
            if inv is None:
                inv = OtherInventoryPeriod(
                    other_product_id=product.id,
                    year=year,
                    month=month,
                    initial_stock=qty,
                    added_stock=0,
                )
                db.add(inv)
            else:
                inv.initial_stock = qty

            imported += 1

        await db.commit()

        file_hash = _compute_file_hash(str(file_path))
        await _log_sync(
            db, original_name, SyncDirection.IMPORT,
            SyncStatus.SUCCESS, imported, file_hash=file_hash,
        )

        return ApiResponse.ok({
            "products_imported": imported,
            "year": year,
            "month": month,
        })

    except Exception as e:
        await _log_sync(
            db, original_name, SyncDirection.IMPORT,
            SyncStatus.FAILED, error=str(e),
        )
        return ApiResponse.fail(f"Import failed: {e}")

    finally:
        file_path.unlink(missing_ok=True)
