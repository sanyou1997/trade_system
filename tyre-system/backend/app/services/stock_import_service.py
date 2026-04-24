import json
import logging
import re
from io import BytesIO
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.inventory import InventoryPeriod
from app.models.other_inventory import OtherInventoryPeriod
from app.models.other_product import OtherProduct
from app.models.phone import Phone
from app.models.phone_inventory import PhoneInventoryPeriod
from app.models.stock_import_log import ImportStatus, StockImportLog
from app.models.tyre import Tyre, TyreCategory
from app.schemas.stock_import import (
    ImportConfirmItem,
    ImportPreviewItem,
    ImportPreviewResult,
    OtherImportConfirmItem,
    OtherImportPreviewItem,
    OtherImportPreviewResult,
    TyreImportConfirmItem,
    TyreImportPreviewItem,
    TyreImportPreviewResult,
)
from app.services.inventory_service import get_inventory
from app.services.other_inventory_service import get_other_inventory
from app.services.phone_inventory_service import get_phone_inventory

logger = logging.getLogger(__name__)


def _normalize(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().lower()


def parse_stock_excel(file_path: str) -> list[dict]:
    """Parse a phone stock Excel file.

    Expected layout:
      A=Package, B=Brand, C=Model, D=Config, E=Quantity
      Row 1 = headers, Row 2+ = data.
      Rows with empty Brand AND Model are skipped (package group headers).
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Excel file has no active sheet")

    rows: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if len(row) < 5:
            continue

        brand = str(row[1].value or "").strip()
        model_name = str(row[2].value or "").strip()
        config = str(row[3].value or "").strip()
        qty_raw = row[4].value

        # Skip package group header rows (no brand/model)
        if not brand and not model_name:
            continue

        quantity = 0
        if qty_raw is not None:
            try:
                quantity = int(float(qty_raw))
            except (ValueError, TypeError):
                pass

        if quantity <= 0:
            continue

        rows.append({
            "row": row_idx,
            "brand": brand,
            "model": model_name,
            "config": config,
            "quantity": quantity,
        })

    wb.close()
    return rows


def _match_phone(phones: list[Phone], brand: str, model_name: str, config: str) -> int | None:
    """Match a phone by brand + model + config (case-insensitive, trimmed)."""
    nb = _normalize(brand)
    nm = _normalize(model_name)
    nc = _normalize(config)

    for p in phones:
        if (
            _normalize(p.brand) == nb
            and _normalize(p.model) == nm
            and _normalize(p.config) == nc
        ):
            return p.id

    # Fallback: brand + model only when config is empty in Excel
    if not nc:
        for p in phones:
            if _normalize(p.brand) == nb and _normalize(p.model) == nm:
                return p.id

    return None


async def preview_import(
    db: AsyncSession,
    file_path: str,
    file_name: str,
    year: int,
    month: int,
) -> ImportPreviewResult:
    """Parse Excel and match against existing phones. Returns preview."""
    parsed_rows = parse_stock_excel(file_path)

    result = await db.execute(select(Phone).order_by(Phone.id))
    all_phones = list(result.scalars().all())

    inv_result = await db.execute(
        select(PhoneInventoryPeriod).where(
            PhoneInventoryPeriod.year == year,
            PhoneInventoryPeriod.month == month,
        )
    )
    inv_by_phone = {inv.phone_id: inv for inv in inv_result.scalars().all()}

    items: list[ImportPreviewItem] = []
    matched_count = 0
    total_qty = 0

    for row_data in parsed_rows:
        phone_id = _match_phone(
            all_phones, row_data["brand"], row_data["model"], row_data["config"]
        )
        matched = phone_id is not None
        current_added = None
        if matched and phone_id in inv_by_phone:
            current_added = inv_by_phone[phone_id].added_stock

        items.append(ImportPreviewItem(
            row_number=row_data["row"],
            brand=row_data["brand"],
            model=row_data["model"],
            config=row_data["config"],
            quantity=row_data["quantity"],
            matched=matched,
            phone_id=phone_id,
            current_added_stock=current_added,
        ))

        if matched:
            matched_count += 1
        total_qty += row_data["quantity"]

    return ImportPreviewResult(
        file_name=file_name,
        total_rows=len(items),
        matched_rows=matched_count,
        unmatched_rows=len(items) - matched_count,
        total_quantity=total_qty,
        items=items,
        all_matched=matched_count == len(items) and len(items) > 0,
    )


async def confirm_import(
    db: AsyncSession,
    year: int,
    month: int,
    file_name: str,
    items: list[ImportConfirmItem],
) -> StockImportLog:
    """Execute the import: add quantities to added_stock and create a log."""
    total_qty = 0
    items_for_log: list[dict] = []

    for item in items:
        result = await db.execute(
            select(PhoneInventoryPeriod).where(
                PhoneInventoryPeriod.phone_id == item.phone_id,
                PhoneInventoryPeriod.year == year,
                PhoneInventoryPeriod.month == month,
            )
        )
        inv = result.scalar_one_or_none()

        if inv is None:
            inv = PhoneInventoryPeriod(
                phone_id=item.phone_id,
                year=year,
                month=month,
                initial_stock=0,
                added_stock=item.quantity,
            )
            db.add(inv)
        else:
            inv.added_stock = inv.added_stock + item.quantity

        total_qty += item.quantity
        items_for_log.append({
            "phone_id": item.phone_id,
            "brand": item.brand,
            "model": item.model,
            "config": item.config,
            "quantity": item.quantity,
        })

    log = StockImportLog(
        product_type="phone",
        year=year,
        month=month,
        file_name=file_name,
        items_json=json.dumps(items_for_log),
        total_quantity=total_qty,
        total_products=len(items),
        status=ImportStatus.ACTIVE,
    )
    db.add(log)
    await db.flush()

    logger.info(
        "Stock import confirmed: %s, %d products, %d total qty for %d/%d",
        file_name, len(items), total_qty, year, month,
    )
    return log


def parse_other_stock_excel(file_path: str) -> list[dict]:
    """Parse an other-product stock Excel file.

    Expected layout:
      A=Name, B=Category, C=Note, D=Suggested Price, E=Quantity.
      Row 1 = headers, Row 2+ = data.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Excel file has no active sheet")

    rows: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if len(row) < 5:
            continue

        name = str(row[0].value or "").strip()
        category = str(row[1].value or "").strip()
        note = str(row[2].value or "").strip()

        suggested_price = 0.0
        if row[3].value is not None:
            try:
                suggested_price = float(row[3].value)
            except (ValueError, TypeError):
                pass

        quantity = 0
        if row[4].value is not None:
            try:
                quantity = int(float(row[4].value))
            except (ValueError, TypeError):
                pass

        if not name or quantity <= 0:
            continue

        rows.append({
            "row": row_idx,
            "name": name,
            "category": category,
            "note": note,
            "suggested_price": suggested_price,
            "quantity": quantity,
        })

    wb.close()
    return rows


def _match_other(products: list[OtherProduct], name: str, category: str) -> int | None:
    nn = _normalize(name)
    nc = _normalize(category)

    for product in products:
        if _normalize(product.name) == nn and _normalize(product.category) == nc:
            return product.id

    if not nc:
        for product in products:
            if _normalize(product.name) == nn:
                return product.id

    return None


async def preview_other_import(
    db: AsyncSession,
    file_path: str,
    file_name: str,
    year: int,
    month: int,
) -> OtherImportPreviewResult:
    parsed_rows = parse_other_stock_excel(file_path)

    result = await db.execute(select(OtherProduct).order_by(OtherProduct.id))
    all_products = list(result.scalars().all())

    inv_result = await db.execute(
        select(OtherInventoryPeriod).where(
            OtherInventoryPeriod.year == year,
            OtherInventoryPeriod.month == month,
        )
    )
    inv_by_product = {
        inv.other_product_id: inv for inv in inv_result.scalars().all()
    }

    items: list[OtherImportPreviewItem] = []
    matched_count = 0
    total_qty = 0

    for row_data in parsed_rows:
        product_id = _match_other(
            all_products, row_data["name"], row_data["category"]
        )
        matched = product_id is not None
        current_added = None
        if matched and product_id in inv_by_product:
            current_added = inv_by_product[product_id].added_stock

        items.append(OtherImportPreviewItem(
            row_number=row_data["row"],
            name=row_data["name"],
            category=row_data["category"],
            note=row_data["note"],
            suggested_price=row_data["suggested_price"],
            quantity=row_data["quantity"],
            matched=matched,
            other_product_id=product_id,
            current_added_stock=current_added,
        ))

        if matched:
            matched_count += 1
        total_qty += row_data["quantity"]

    return OtherImportPreviewResult(
        file_name=file_name,
        total_rows=len(items),
        matched_rows=matched_count,
        unmatched_rows=len(items) - matched_count,
        total_quantity=total_qty,
        items=items,
        all_matched=matched_count == len(items) and len(items) > 0,
    )


async def confirm_other_import(
    db: AsyncSession,
    year: int,
    month: int,
    file_name: str,
    items: list[OtherImportConfirmItem],
) -> StockImportLog:
    total_qty = 0
    items_for_log: list[dict] = []

    for item in items:
        product_id = item.other_product_id

        if item.create_new and product_id is None:
            new_product = OtherProduct(
                name=item.name,
                category=item.category or None,
                note=item.note or None,
                suggested_price=item.suggested_price,
                cost=0.0,
            )
            db.add(new_product)
            await db.flush()
            product_id = new_product.id

        if product_id is None:
            continue

        result = await db.execute(
            select(OtherInventoryPeriod).where(
                OtherInventoryPeriod.other_product_id == product_id,
                OtherInventoryPeriod.year == year,
                OtherInventoryPeriod.month == month,
            )
        )
        inv = result.scalar_one_or_none()

        if inv is None:
            inv = OtherInventoryPeriod(
                other_product_id=product_id,
                year=year,
                month=month,
                initial_stock=0,
                added_stock=item.quantity,
            )
            db.add(inv)
        else:
            inv.added_stock = inv.added_stock + item.quantity

        total_qty += item.quantity
        items_for_log.append({
            "other_product_id": product_id,
            "name": item.name,
            "category": item.category,
            "quantity": item.quantity,
            "created_new": item.create_new,
        })

    log = StockImportLog(
        product_type="other",
        year=year,
        month=month,
        file_name=file_name,
        items_json=json.dumps(items_for_log),
        total_quantity=total_qty,
        total_products=len(items_for_log),
        status=ImportStatus.ACTIVE,
    )
    db.add(log)
    await db.flush()

    logger.info(
        "Other stock import confirmed: %s, %d products, %d total qty for %d/%d",
        file_name, len(items_for_log), total_qty, year, month,
    )
    return log


async def revert_import(
    db: AsyncSession,
    log_id: int,
) -> StockImportLog:
    """Undo an import by subtracting quantities from added_stock."""
    result = await db.execute(
        select(StockImportLog).where(StockImportLog.id == log_id)
    )
    log = result.scalar_one_or_none()

    if log is None:
        raise ValueError(f"Import log {log_id} not found")
    if log.status == ImportStatus.REVERTED:
        raise ValueError(f"Import log {log_id} has already been reverted")

    items = json.loads(log.items_json)

    if log.product_type == "tyre":
        for item in items:
            inv_result = await db.execute(
                select(InventoryPeriod).where(
                    InventoryPeriod.tyre_id == item["tyre_id"],
                    InventoryPeriod.year == log.year,
                    InventoryPeriod.month == log.month,
                )
            )
            inv = inv_result.scalar_one_or_none()
            if inv is not None:
                inv.added_stock = max(0, inv.added_stock - item["quantity"])
    elif log.product_type == "other":
        for item in items:
            inv_result = await db.execute(
                select(OtherInventoryPeriod).where(
                    OtherInventoryPeriod.other_product_id == item["other_product_id"],
                    OtherInventoryPeriod.year == log.year,
                    OtherInventoryPeriod.month == log.month,
                )
            )
            inv = inv_result.scalar_one_or_none()
            if inv is not None:
                inv.added_stock = max(0, inv.added_stock - item["quantity"])
    else:
        for item in items:
            inv_result = await db.execute(
                select(PhoneInventoryPeriod).where(
                    PhoneInventoryPeriod.phone_id == item["phone_id"],
                    PhoneInventoryPeriod.year == log.year,
                    PhoneInventoryPeriod.month == log.month,
                )
            )
            inv = inv_result.scalar_one_or_none()
            if inv is not None:
                inv.added_stock = max(0, inv.added_stock - item["quantity"])

    log.status = ImportStatus.REVERTED
    log.reverted_at = datetime.now(timezone.utc)
    await db.flush()

    logger.info("Stock import %d reverted", log_id)
    return log


async def get_import_history(
    db: AsyncSession,
    product_type: str = "phone",
    limit: int = 50,
) -> list[StockImportLog]:
    """Get recent import logs."""
    result = await db.execute(
        select(StockImportLog)
        .where(StockImportLog.product_type == product_type)
        .order_by(StockImportLog.created_at.desc())
        .limit(limit)
    )
    return list(result.scalars().all())


def _format_stock_sheet(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 14)


async def export_stock_workbook(
    db: AsyncSession,
    product_type: str,
    year: int,
    month: int,
) -> BytesIO:
    """Export current remaining stock in the same shape accepted by stock import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    if product_type == "tyre":
        _format_stock_sheet(ws, [
            "Size",
            "Type",
            "Brand",
            "Pattern",
            "LI & SR",
            "Tyre COST",
            "After Delivery COST",
            "QTY",
            "Suggested Price",
        ])
        for item in await get_inventory(db, year, month):
            ws.append([
                item["size"],
                item["type"],
                item["brand"] or "",
                item["pattern"] or "",
                item.get("li_sr") or "",
                item["tyre_cost"],
                "",
                item["remaining_stock"],
                item["suggested_price"],
            ])
    elif product_type == "phone":
        _format_stock_sheet(ws, ["Package", "Brand", "Model", "Config", "Quantity"])
        for item in await get_phone_inventory(db, year, month):
            ws.append([
                "",
                item["brand"],
                item["model"],
                item["config"] or "",
                item["remaining_stock"],
            ])
    elif product_type == "other":
        _format_stock_sheet(ws, [
            "Name",
            "Category",
            "Note",
            "Suggested Price",
            "Quantity",
        ])
        for item in await get_other_inventory(db, year, month):
            ws.append([
                item["name"],
                item["category"] or "",
                item["note"] or "",
                item["suggested_price"],
                item["remaining_stock"],
            ])
    else:
        raise ValueError(f"Unsupported product_type: {product_type}")

    stream = BytesIO()
    wb.save(stream)
    wb.close()
    stream.seek(0)
    return stream


# ---------------------------------------------------------------------------
# Tyre stock import
# ---------------------------------------------------------------------------

def _normalize_size(size: str) -> str:
    """Normalize tyre size for comparison.

    Strips whitespace, lowercases, and normalizes common variations:
    '205/65 R15 ' -> '205/65r15', '195R14C' -> '195r14c'
    """
    s = size.strip().lower()
    # Remove spaces around slashes and R
    s = re.sub(r"\s+", "", s)
    return s


def _match_tyre(tyres: list[Tyre], size: str, brand: str) -> int | None:
    """Match a tyre by normalized size + brand (case-insensitive)."""
    ns = _normalize_size(size)
    nb = _normalize(brand)

    for t in tyres:
        if _normalize_size(t.size) == ns and _normalize(t.brand or "") == nb:
            return t.id
    return None


def parse_tyre_stock_excel(file_path: str) -> list[dict]:
    """Parse a tyre stock Excel file.

    Expected layout:
      A=Size, B=Type, C=Brand, D=Pattern, E=LI & SR,
      F=Tyre COST, G=After Delivery COST (formula), H=QTY,
      I=Suggested Price (formula)
      Row 1 = headers, Row 2+ = data.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Excel file has no active sheet")

    rows: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if len(row) < 8:
            continue

        size = str(row[0].value or "").strip()
        type_ = str(row[1].value or "").strip()
        brand = str(row[2].value or "").strip()
        pattern = str(row[3].value or "").strip()
        li_sr = str(row[4].value or "").strip()

        # Column F: tyre cost (CNY)
        tyre_cost = 0.0
        if row[5].value is not None:
            try:
                tyre_cost = float(row[5].value)
            except (ValueError, TypeError):
                pass

        # Column H: quantity
        qty_raw = row[7].value
        quantity = 0
        if qty_raw is not None:
            try:
                quantity = int(float(qty_raw))
            except (ValueError, TypeError):
                pass

        # Column I: suggested price (formula, read via data_only)
        suggested_price = 0.0
        if len(row) > 8 and row[8].value is not None:
            try:
                suggested_price = float(row[8].value)
            except (ValueError, TypeError):
                pass

        # Skip rows with no size or no quantity
        if not size or quantity <= 0:
            continue

        rows.append({
            "row": row_idx,
            "size": size,
            "type_": type_,
            "brand": brand,
            "pattern": pattern,
            "li_sr": li_sr,
            "tyre_cost": tyre_cost,
            "suggested_price": suggested_price,
            "quantity": quantity,
        })

    wb.close()
    return rows


async def preview_tyre_import(
    db: AsyncSession,
    file_path: str,
    file_name: str,
    year: int,
    month: int,
) -> TyreImportPreviewResult:
    """Parse tyre Excel and match against existing tyres. Returns preview."""
    parsed_rows = parse_tyre_stock_excel(file_path)

    result = await db.execute(select(Tyre).order_by(Tyre.id))
    all_tyres = list(result.scalars().all())

    inv_result = await db.execute(
        select(InventoryPeriod).where(
            InventoryPeriod.year == year,
            InventoryPeriod.month == month,
        )
    )
    inv_by_tyre = {inv.tyre_id: inv for inv in inv_result.scalars().all()}

    items: list[TyreImportPreviewItem] = []
    matched_count = 0
    total_qty = 0

    for row_data in parsed_rows:
        tyre_id = _match_tyre(all_tyres, row_data["size"], row_data["brand"])
        matched = tyre_id is not None
        current_added = None
        if matched and tyre_id in inv_by_tyre:
            current_added = inv_by_tyre[tyre_id].added_stock

        items.append(TyreImportPreviewItem(
            row_number=row_data["row"],
            size=row_data["size"],
            type_=row_data["type_"],
            brand=row_data["brand"],
            pattern=row_data["pattern"],
            li_sr=row_data["li_sr"],
            tyre_cost=row_data["tyre_cost"],
            suggested_price=row_data["suggested_price"],
            quantity=row_data["quantity"],
            matched=matched,
            tyre_id=tyre_id,
            current_added_stock=current_added,
        ))

        if matched:
            matched_count += 1
        total_qty += row_data["quantity"]

    return TyreImportPreviewResult(
        file_name=file_name,
        total_rows=len(items),
        matched_rows=matched_count,
        unmatched_rows=len(items) - matched_count,
        total_quantity=total_qty,
        items=items,
        all_matched=matched_count == len(items) and len(items) > 0,
    )


def _category_from_brand(brand: str) -> TyreCategory:
    """Infer tyre category from brand name."""
    if not brand:
        return TyreCategory.BRANDLESS_NEW
    return TyreCategory.BRANDED_NEW


async def confirm_tyre_import(
    db: AsyncSession,
    year: int,
    month: int,
    file_name: str,
    items: list[TyreImportConfirmItem],
) -> StockImportLog:
    """Execute tyre import: create new products if needed, add stock, log."""
    total_qty = 0
    items_for_log: list[dict] = []

    for item in items:
        tyre_id = item.tyre_id

        # Create new tyre product if requested
        if item.create_new and tyre_id is None:
            category = TyreCategory(item.category) if item.category else _category_from_brand(item.brand)
            new_tyre = Tyre(
                size=item.size,
                type_=item.type_,
                brand=item.brand or None,
                pattern=item.pattern or None,
                li_sr=item.li_sr or None,
                tyre_cost=item.tyre_cost,
                suggested_price=item.suggested_price,
                category=category,
            )
            db.add(new_tyre)
            await db.flush()  # get the new ID
            tyre_id = new_tyre.id

        if tyre_id is None:
            continue  # skip unmatched items not marked for creation

        # Update or create inventory period
        result = await db.execute(
            select(InventoryPeriod).where(
                InventoryPeriod.tyre_id == tyre_id,
                InventoryPeriod.year == year,
                InventoryPeriod.month == month,
            )
        )
        inv = result.scalar_one_or_none()

        if inv is None:
            inv = InventoryPeriod(
                tyre_id=tyre_id,
                year=year,
                month=month,
                initial_stock=0,
                added_stock=item.quantity,
            )
            db.add(inv)
        else:
            inv.added_stock = inv.added_stock + item.quantity

        total_qty += item.quantity
        items_for_log.append({
            "tyre_id": tyre_id,
            "size": item.size,
            "brand": item.brand,
            "quantity": item.quantity,
            "created_new": item.create_new,
        })

    log = StockImportLog(
        product_type="tyre",
        year=year,
        month=month,
        file_name=file_name,
        items_json=json.dumps(items_for_log),
        total_quantity=total_qty,
        total_products=len(items_for_log),
        status=ImportStatus.ACTIVE,
    )
    db.add(log)
    await db.flush()

    logger.info(
        "Tyre stock import confirmed: %s, %d products, %d total qty for %d/%d",
        file_name, len(items_for_log), total_qty, year, month,
    )
    return log
