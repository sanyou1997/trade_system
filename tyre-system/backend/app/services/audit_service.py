from datetime import date

from sqlalchemy import func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_account import AuditAccount
from app.models.audit_balance_override import AuditBalanceOverride
from app.models.audit_transaction import AuditTransaction, TransactionType
from app.models.phone_sale import PhoneSale
from app.models.sale import PaymentMethod, Sale
from app.schemas.audit import (
    AccountBalanceResponse,
    AccountCreate,
    AccountUpdate,
    ExchangeCreate,
    ExpenseCreate,
    IncomeCreate,
    RevenueBreakdown,
    TransactionFilter,
    TransferCreate,
)


# --- Account CRUD ---


async def create_account(db: AsyncSession, data: AccountCreate) -> AuditAccount:
    if data.is_default:
        await _clear_defaults(db)
    account = AuditAccount(
        name=data.name,
        description=data.description,
        initial_balance=data.initial_balance,
        is_default=data.is_default,
    )
    db.add(account)
    await db.flush()
    await db.refresh(account)
    return account


async def update_account(
    db: AsyncSession, account_id: int, data: AccountUpdate
) -> AuditAccount | None:
    result = await db.execute(
        select(AuditAccount).where(AuditAccount.id == account_id)
    )
    account = result.scalar_one_or_none()
    if account is None:
        return None
    if data.is_default:
        await _clear_defaults(db)
    if data.name is not None:
        account.name = data.name
    if data.description is not None:
        account.description = data.description
    if data.initial_balance is not None:
        account.initial_balance = data.initial_balance
    if data.is_default is not None:
        account.is_default = data.is_default
    await db.flush()
    await db.refresh(account)
    return account


async def get_accounts(db: AsyncSession) -> list[AuditAccount]:
    result = await db.execute(
        select(AuditAccount).order_by(
            AuditAccount.is_default.desc(), AuditAccount.id
        )
    )
    return list(result.scalars().all())


async def delete_account(db: AsyncSession, account_id: int) -> bool:
    result = await db.execute(
        select(AuditAccount).where(AuditAccount.id == account_id)
    )
    account = result.scalar_one_or_none()
    if account is None:
        return False
    if account.is_default:
        raise ValueError("Cannot delete the default account")

    # Check cumulative balance is zero
    balance = await _calculate_account_balance(db, account_id, None, None)
    if abs(balance) > 0.01:
        raise ValueError(
            f"Cannot delete account with non-zero balance ({balance:.0f} MWK)"
        )

    await db.delete(account)
    await db.flush()
    return True


async def _clear_defaults(db: AsyncSession) -> None:
    await db.execute(
        update(AuditAccount)
        .where(AuditAccount.is_default == True)  # noqa: E712
        .values(is_default=False)
    )


# --- Transaction CRUD ---


async def create_expense(db: AsyncSession, data: ExpenseCreate) -> AuditTransaction:
    txn = AuditTransaction(
        transaction_type=TransactionType.EXPENSE,
        transaction_date=data.transaction_date,
        description=data.description,
        amount_mwk=data.amount_mwk,
        account_id=data.account_id,
        receipt_info=data.receipt_info,
        note=data.note,
    )
    db.add(txn)
    await db.flush()
    await db.refresh(txn)
    return txn


async def create_transfer(db: AsyncSession, data: TransferCreate) -> AuditTransaction:
    if data.from_account_id == data.to_account_id:
        raise ValueError("Cannot transfer to the same account")
    txn = AuditTransaction(
        transaction_type=TransactionType.TRANSFER,
        transaction_date=data.transaction_date,
        description=data.description or "Transfer",
        amount_mwk=data.amount_mwk,
        from_account_id=data.from_account_id,
        to_account_id=data.to_account_id,
        note=data.note,
    )
    db.add(txn)
    await db.flush()
    await db.refresh(txn)
    return txn


async def create_exchange(db: AsyncSession, data: ExchangeCreate) -> AuditTransaction:
    txn = AuditTransaction(
        transaction_type=TransactionType.EXCHANGE,
        transaction_date=data.transaction_date,
        description=data.description or f"Exchange at rate {data.exchange_rate}",
        amount_mwk=data.amount_mwk,
        exchange_rate=data.exchange_rate,
        amount_cny=data.amount_cny,
        account_id=data.account_id,
        note=data.note,
    )
    db.add(txn)
    await db.flush()
    await db.refresh(txn)
    return txn


async def create_income(db: AsyncSession, data: IncomeCreate) -> AuditTransaction:
    txn = AuditTransaction(
        transaction_type=TransactionType.INCOME,
        transaction_date=data.transaction_date,
        description=data.description,
        amount_mwk=data.amount_mwk,
        account_id=data.account_id,
        note=data.note,
    )
    db.add(txn)
    await db.flush()
    await db.refresh(txn)
    return txn


async def get_transactions(
    db: AsyncSession, filters: TransactionFilter
) -> tuple[list[AuditTransaction], int]:
    query = select(AuditTransaction)
    count_query = select(func.count(AuditTransaction.id))

    conditions = []
    if filters.year is not None and filters.month is not None:
        conditions.append(
            func.extract("year", AuditTransaction.transaction_date) == filters.year
        )
        conditions.append(
            func.extract("month", AuditTransaction.transaction_date) == filters.month
        )
    if filters.transaction_type is not None:
        conditions.append(
            AuditTransaction.transaction_type == filters.transaction_type
        )
    if filters.account_id is not None:
        conditions.append(
            or_(
                AuditTransaction.account_id == filters.account_id,
                AuditTransaction.from_account_id == filters.account_id,
                AuditTransaction.to_account_id == filters.account_id,
            )
        )

    if conditions:
        query = query.where(*conditions)
        count_query = count_query.where(*conditions)

    total_result = await db.execute(count_query)
    total = total_result.scalar()

    offset = (filters.page - 1) * filters.limit
    query = query.order_by(
        AuditTransaction.transaction_date.desc(), AuditTransaction.id.desc()
    )
    query = query.offset(offset).limit(filters.limit)

    result = await db.execute(query)
    return list(result.scalars().all()), total


async def delete_transaction(db: AsyncSession, txn_id: int) -> bool:
    result = await db.execute(
        select(AuditTransaction).where(AuditTransaction.id == txn_id)
    )
    txn = result.scalar_one_or_none()
    if txn is None:
        return False
    await db.delete(txn)
    await db.flush()
    return True


async def upload_receipt_image(
    db: AsyncSession, txn_id: int, filename: str
) -> AuditTransaction | None:
    result = await db.execute(
        select(AuditTransaction).where(AuditTransaction.id == txn_id)
    )
    txn = result.scalar_one_or_none()
    if txn is None:
        return None
    txn.receipt_image = filename
    await db.flush()
    await db.refresh(txn)
    return txn


# --- Revenue Aggregation ---


async def get_revenue_breakdown(
    db: AsyncSession, year: int, month: int
) -> RevenueBreakdown:
    breakdown = RevenueBreakdown()

    for method, field in [
        (PaymentMethod.CASH, "tyre_cash"),
        (PaymentMethod.MUKURU, "tyre_mukuru"),
        (PaymentMethod.CARD, "tyre_card"),
    ]:
        result = await db.execute(
            select(func.coalesce(func.sum(Sale.total), 0)).where(
                func.extract("year", Sale.sale_date) == year,
                func.extract("month", Sale.sale_date) == month,
                Sale.payment_method == method,
            )
        )
        setattr(breakdown, field, float(result.scalar()))

    for method, field in [
        (PaymentMethod.CASH, "phone_cash"),
        (PaymentMethod.MUKURU, "phone_mukuru"),
        (PaymentMethod.CARD, "phone_card"),
    ]:
        result = await db.execute(
            select(func.coalesce(func.sum(PhoneSale.total), 0)).where(
                func.extract("year", PhoneSale.sale_date) == year,
                func.extract("month", PhoneSale.sale_date) == month,
                PhoneSale.payment_method == method,
            )
        )
        setattr(breakdown, field, float(result.scalar()))

    breakdown.tyre_total = (
        breakdown.tyre_cash + breakdown.tyre_mukuru + breakdown.tyre_card
    )
    breakdown.phone_total = (
        breakdown.phone_cash + breakdown.phone_mukuru + breakdown.phone_card
    )
    breakdown.grand_total = breakdown.tyre_total + breakdown.phone_total
    return breakdown


# --- Cumulative Revenue (all months up to target) ---


async def _cumulative_revenue(
    db: AsyncSession,
    up_to_year: int | None,
    up_to_month: int | None,
) -> float:
    """Sum of all tyre + phone sales revenue up to given year/month."""
    total = 0.0

    for model in [Sale, PhoneSale]:
        query = select(func.coalesce(func.sum(model.total), 0))
        if up_to_year is not None and up_to_month is not None:
            query = query.where(
                (func.extract("year", model.sale_date) < up_to_year)
                | (
                    (func.extract("year", model.sale_date) == up_to_year)
                    & (func.extract("month", model.sale_date) <= up_to_month)
                )
            )
        result = await db.execute(query)
        total += float(result.scalar())

    return total


# --- Balance Calculation ---


def _month_lte_condition(date_col, year: int, month: int):
    """SQLAlchemy condition: date_col's (year, month) <= (year, month)."""
    return (func.extract("year", date_col) < year) | (
        (func.extract("year", date_col) == year)
        & (func.extract("month", date_col) <= month)
    )


def _month_eq_condition(date_col, year: int, month: int):
    """SQLAlchemy condition: date_col's (year, month) == (year, month)."""
    return (func.extract("year", date_col) == year) & (
        func.extract("month", date_col) == month
    )


async def _sum_transactions(
    db: AsyncSession,
    txn_type: TransactionType,
    account_field: str,
    account_id: int,
    year: int | None,
    month: int | None,
    *,
    exact_month: bool = False,
) -> float:
    """Sum amount_mwk for a transaction type and account field.

    If exact_month=True, sum only transactions in the exact year/month.
    Otherwise, sum cumulatively up to and including year/month.
    """
    col = getattr(AuditTransaction, account_field)
    query = select(func.coalesce(func.sum(AuditTransaction.amount_mwk), 0)).where(
        AuditTransaction.transaction_type == txn_type,
        col == account_id,
    )
    if year is not None and month is not None:
        if exact_month:
            query = query.where(
                _month_eq_condition(AuditTransaction.transaction_date, year, month)
            )
        else:
            query = query.where(
                _month_lte_condition(AuditTransaction.transaction_date, year, month)
            )
    result = await db.execute(query)
    return float(result.scalar())


async def _calculate_account_balance(
    db: AsyncSession,
    account_id: int,
    year: int | None,
    month: int | None,
) -> float:
    """Calculate cumulative balance for an account, excluding auto-revenue."""
    expenses = await _sum_transactions(
        db, TransactionType.EXPENSE, "account_id", account_id, year, month
    )
    exchanges = await _sum_transactions(
        db, TransactionType.EXCHANGE, "account_id", account_id, year, month
    )
    income = await _sum_transactions(
        db, TransactionType.INCOME, "account_id", account_id, year, month
    )
    transfers_in = await _sum_transactions(
        db, TransactionType.TRANSFER, "to_account_id", account_id, year, month
    )
    transfers_out = await _sum_transactions(
        db, TransactionType.TRANSFER, "from_account_id", account_id, year, month
    )
    return income - expenses - exchanges + transfers_in - transfers_out


def _prev_month(year: int, month: int) -> tuple[int, int]:
    """Return (year, month) for the previous month."""
    if month == 1:
        return year - 1, 12
    return year, month - 1


async def _month_revenue(
    db: AsyncSession, year: int, month: int
) -> float:
    """Sum of all tyre + phone sales revenue for a single month."""
    total = 0.0
    for model in [Sale, PhoneSale]:
        result = await db.execute(
            select(func.coalesce(func.sum(model.total), 0)).where(
                func.extract("year", model.sale_date) == year,
                func.extract("month", model.sale_date) == month,
            )
        )
        total += float(result.scalar())
    return total


async def _cumulative_balance_through(
    db: AsyncSession, account_id: int, is_default: bool,
    year: int, month: int, initial_balance: float,
) -> float:
    """Full cumulative balance for an account through the given year/month."""
    revenue = 0.0
    if is_default:
        revenue = await _cumulative_revenue(db, year, month)
    txn_balance = await _calculate_account_balance(
        db, account_id, year, month
    )
    return initial_balance + revenue + txn_balance


async def _get_override(
    db: AsyncSession, account_id: int, year: int, month: int
) -> AuditBalanceOverride | None:
    result = await db.execute(
        select(AuditBalanceOverride).where(
            AuditBalanceOverride.account_id == account_id,
            AuditBalanceOverride.year == year,
            AuditBalanceOverride.month == month,
        )
    )
    return result.scalar_one_or_none()


async def set_balance_override(
    db: AsyncSession, account_id: int, year: int, month: int, value: float
) -> None:
    """Set or update the initial balance override for a specific account+month."""
    existing = await _get_override(db, account_id, year, month)
    if existing is not None:
        existing.override_balance = value
    else:
        db.add(AuditBalanceOverride(
            account_id=account_id, year=year, month=month, override_balance=value,
        ))
    await db.flush()


async def clear_balance_override(
    db: AsyncSession, account_id: int, year: int, month: int
) -> bool:
    """Remove the initial balance override, reverting to auto-calculated."""
    existing = await _get_override(db, account_id, year, month)
    if existing is None:
        return False
    await db.delete(existing)
    await db.flush()
    return True


async def get_account_balances(
    db: AsyncSession, year: int, month: int
) -> list[AccountBalanceResponse]:
    accounts = await get_accounts(db)
    prev_y, prev_m = _prev_month(year, month)

    results = []
    for acct in accounts:
        # Check for manual override first
        override = await _get_override(db, acct.id, year, month)
        has_override = override is not None

        if has_override:
            prev_bal = override.override_balance
        else:
            # Auto-calculated: cumulative balance through previous month
            prev_bal = await _cumulative_balance_through(
                db, acct.id, acct.is_default, prev_y, prev_m, acct.initial_balance
            )

        # Current month only: revenue
        auto_rev = (
            await _month_revenue(db, year, month)
            if acct.is_default
            else 0.0
        )

        # Current month only: transactions
        expenses = await _sum_transactions(
            db, TransactionType.EXPENSE, "account_id", acct.id,
            year, month, exact_month=True,
        )
        exchanges = await _sum_transactions(
            db, TransactionType.EXCHANGE, "account_id", acct.id,
            year, month, exact_month=True,
        )
        manual_income = await _sum_transactions(
            db, TransactionType.INCOME, "account_id", acct.id,
            year, month, exact_month=True,
        )
        transfers_in = await _sum_transactions(
            db, TransactionType.TRANSFER, "to_account_id", acct.id,
            year, month, exact_month=True,
        )
        transfers_out = await _sum_transactions(
            db, TransactionType.TRANSFER, "from_account_id", acct.id,
            year, month, exact_month=True,
        )

        calculated_balance = (
            prev_bal
            + auto_rev
            + manual_income
            - expenses
            - exchanges
            + transfers_in
            - transfers_out
        )

        results.append(
            AccountBalanceResponse(
                id=acct.id,
                name=acct.name,
                description=acct.description,
                initial_balance=acct.initial_balance,
                is_default=acct.is_default,
                prev_balance=round(prev_bal, 2),
                has_override=has_override,
                auto_revenue=auto_rev,
                manual_income=manual_income,
                total_expenses=expenses,
                total_exchanges=exchanges,
                transfers_in=transfers_in,
                transfers_out=transfers_out,
                calculated_balance=round(calculated_balance, 2),
            )
        )

    return results


# --- Excel Import ---


async def import_from_audit_excel(
    db: AsyncSession, file_path: str, default_account_id: int
) -> dict:
    """Parse the Audit_2026.xlsx file and import expense transactions."""
    import openpyxl
    from datetime import datetime as dt

    wb = openpyxl.load_workbook(file_path, data_only=True)
    cutoff_date = date(2026, 2, 16)

    imported_expenses = 0
    imported_exchanges = 0
    skipped = 0
    errors: list[str] = []

    # Sheets to process: monthly expense sheets (not Profit Share)
    target_sheets = []
    for name in wb.sheetnames:
        if "月" in name and "Profit" not in name:
            target_sheets.append(name)

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        for row_idx in range(2, ws.max_row + 1):
            # Column J has category indicators (Total, Exchange, Expense) - skip summary rows
            col_j = ws.cell(row=row_idx, column=10).value
            if col_j is not None and str(col_j).strip():
                continue

            raw_date = ws.cell(row=row_idx, column=1).value
            description = ws.cell(row=row_idx, column=2).value
            amount = ws.cell(row=row_idx, column=6).value

            if not description or not amount:
                continue

            # Parse date
            txn_date = None
            if isinstance(raw_date, (int, float)):
                try:
                    txn_date = dt.fromordinal(
                        dt(1899, 12, 30).toordinal() + int(raw_date)
                    ).date()
                except (ValueError, OverflowError):
                    pass
            elif isinstance(raw_date, dt):
                txn_date = raw_date.date()
            elif isinstance(raw_date, date):
                txn_date = raw_date
            elif isinstance(raw_date, str):
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        txn_date = dt.strptime(raw_date.strip(), fmt).date()
                        break
                    except ValueError:
                        continue

            if txn_date is None:
                errors.append(f"Sheet '{sheet_name}' row {row_idx}: invalid date")
                continue

            if txn_date > cutoff_date:
                continue

            amount_val = float(amount)
            if amount_val <= 0:
                continue

            desc_str = str(description).strip()
            receipt_info = ws.cell(row=row_idx, column=8).value
            receipt_str = str(receipt_info).strip() if receipt_info else None

            # Duplicate detection
            existing = await db.execute(
                select(func.count(AuditTransaction.id)).where(
                    AuditTransaction.transaction_date == txn_date,
                    AuditTransaction.description == desc_str,
                    AuditTransaction.amount_mwk == amount_val,
                )
            )
            if existing.scalar() > 0:
                skipped += 1
                continue

            # Determine type: if description contains "exchange", create as exchange
            desc_lower = desc_str.lower()
            if "exchange" in desc_lower:
                txn = AuditTransaction(
                    transaction_type=TransactionType.EXCHANGE,
                    transaction_date=txn_date,
                    description=desc_str,
                    amount_mwk=amount_val,
                    account_id=default_account_id,
                    receipt_info=receipt_str,
                )
                db.add(txn)
                imported_exchanges += 1
            else:
                txn = AuditTransaction(
                    transaction_type=TransactionType.EXPENSE,
                    transaction_date=txn_date,
                    description=desc_str,
                    amount_mwk=amount_val,
                    account_id=default_account_id,
                    receipt_info=receipt_str,
                )
                db.add(txn)
                imported_expenses += 1

    await db.flush()
    return {
        "expenses_imported": imported_expenses,
        "exchanges_imported": imported_exchanges,
        "skipped": skipped,
        "errors": errors,
    }
